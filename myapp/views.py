from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from .forms import  TaskForm, UserProfileForm, ProjectMemberForm, ClientForm, MailForm
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate, login,logout
from django.http import JsonResponse, FileResponse, Http404
from django.views import View
from rest_framework.exceptions import AuthenticationFailed
from django.contrib import messages
from django.http import Http404,HttpResponse,HttpResponseForbidden
from django.shortcuts import redirect
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.clickjacking import xframe_options_exempt
import os
from django.http import JsonResponse
from collections import defaultdict
import json
from .serializers import RegisterSerializer, LoginSerializer, UserProfileSerializer,TaskSerializer
from rest_framework.authtoken.models import Token
from django.conf import settings
from django.core.mail import send_mail, EmailMessage
from django.urls import reverse
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.auth.hashers import make_password
from django.conf import settings
import mimetypes
from django.contrib.auth.hashers import check_password
from django.core.files.storage import default_storage
from django.db.models import Count, Case, When, Value, IntegerField
from django.utils import timezone
from rest_framework import status
from django.db.models import Q
from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework_simplejwt.exceptions import TokenError, InvalidToken
from django.views.decorators.http import require_POST
import json
from datetime import datetime
from django.contrib.contenttypes.models import ContentType
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import io
from django.db import models
import re


class DualAuthentication:
    def authenticate(self, request):
        session_auth = SessionAuthentication()
        jwt_auth = JWTAuthentication()

        session_result = session_auth.authenticate(request)
        jwt_result = jwt_auth.authenticate(request)

        if not session_result or not jwt_result:
            raise AuthenticationFailed("Both session and JWT authentication are required.")

        return jwt_result  # or session_result (user should be same)

    def authenticate_header(self, request):
        # Combine both headers for clarity, or just return one
        return 'Bearer, session'
class RegisterAPIView(View):

    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        role = getattr(request.user.userprofile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_REGISTER_PAGE_VISIT",
            f"{role.title()} {request.user.username} visited the Registration page."
        )

        print('register get request')
        last_profile = UserProfile.objects.order_by('-id').first()

        if last_profile and last_profile.empid.isdigit():
            next_empid = str(int(last_profile.empid) + 1)
        else:
            next_empid = "1001"

        return render(request, 'registration.html', {
            'next_empid': next_empid
        })

    def post(self, request):
        print('inside post registration')
        try:
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            role = request.POST.get('role')
            phone = request.POST.get('phone')
            empid = request.POST.get('empid')

            data = {
                'empname': username,
                'email': email,
                'password': password,
                'role': role,
                'phone': phone,
                'empid': empid
            }

            print("Registration data:", data)

            serializer = RegisterSerializer(data=data)

            if serializer.is_valid():
                try:
                    user_profile = serializer.save()

                    create_activity_log(
                        user_profile.user,
                        'USER_REGISTERED',
                        f'New user {user_profile.user.username} registered.'
                    )

                    # Notify all managers
                    managers = User.objects.filter(userprofile__role='Manager')
                    for manager in managers:
                        create_notification(
                            user=manager,
                            title='New User Registered',
                            message=f'User "{user_profile.user.username}" has joined as {user_profile.role}.',
                            notification_type='User',
                            link=f'/users/{user_profile.user.id}/view/'
                        )

                    messages.success(request, 'Registration successful! Please login.')

                except Exception as e:
                    print(f"Error saving user: {str(e)}")
                    messages.error(request, f'Registration failed: {str(e)}')

            else:
                print("Serializer errors:", serializer.errors)
                error_messages = [
                    f"{field}: {error}"
                    for field, errors in serializer.errors.items()
                    for error in errors
                ]
                messages.error(request, 'Registration failed: ' + ', '.join(error_messages))

        except Exception as e:
            print(f"Registration error: {str(e)}")
            messages.error(request, f'Registration error: {str(e)}')

        # Recalculate next employee ID
        last_profile = UserProfile.objects.order_by('-id').first()
        if last_profile and last_profile.empid.isdigit():
            next_empid = str(int(last_profile.empid) + 1)
        else:
            next_empid = "1001"

        return render(request, 'registration.html', {
            'next_empid': next_empid,
            'form_data': request.POST
        })



class LandingView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def get(self, request):
        return render(request, 'landing.html')  
    

class GetStarted(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def get(self, request):
        return render(request, 'get_started.html')


class LoginView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]
 
    def get(self, request):
        return render(request, 'login.html')  # Render the login page
 
    def post(self, request):
        # --- Robust parsing ---
        try:
            # DRF handles request.data, but this ensures fallback if empty
            data = request.data or json.loads(request.body.decode('utf-8'))
        except Exception:
            data = {}

        serializer = LoginSerializer(data=data)

        if not serializer.is_valid():
            return JsonResponse({'error': serializer.errors}, status=400)

        empid = serializer.validated_data.get('employeeId')
        password = serializer.validated_data.get('password')

        try:
            user_profile = None
            user = None
            
            try:
                user_profile = UserProfile.objects.get(empid=empid)
                user = user_profile.user

            except UserProfile.DoesNotExist:
                superuser = User.objects.filter(is_superuser=True).first()

                if empid == "1000" and superuser:
                    user_profile = UserProfile.objects.create(
                        user=superuser,
                        empid="1000",
                        name=superuser.username,
                        email=superuser.email or "admin@example.com",
                        role="Manager",
                        phone="961XXXXXXX",
                        password=superuser.password
                    )
                    user = superuser
                else:
                    return JsonResponse({'error': 'Invalid Employee ID'}, status=400)

            if not user_profile or not user:
                return JsonResponse({'error': 'User profile not found'}, status=400)

            authenticated_user = authenticate(request, username=user.username, password=password)

            if authenticated_user is not None:
                login(request, authenticated_user)

                role = user_profile.role
                log_type = f"{role.replace(' ', '_').upper()}_LOGGED_IN"
                message = f"{role} {authenticated_user.username} logged in successfully."

                create_activity_log(authenticated_user, log_type, message)
                refresh = RefreshToken.for_user(authenticated_user)

                return JsonResponse({
                    'message': f'Welcome {authenticated_user.username}!',
                    'access_token': str(refresh.access_token),
                    'refresh_token': str(refresh),
                    'username': authenticated_user.username,
                    'role': role,
                    'empid': user_profile.empid,
                }, status=200)

            else:
                return JsonResponse({'error': 'Invalid password'}, status=401)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
       
class UserProfileAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            user_profile = UserProfile.objects.get(user=request.user)
            serializer = UserProfileSerializer(user_profile)
            return JsonResponse({'user':serializer.data}, status=status.HTTP_200_OK)
        except UserProfile.DoesNotExist:
            return JsonResponse({'error': 'User profile not found'}, status=status.HTTP_404_NOT_FOUND)

class DashboardView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')

        # Log the dashboard visit with username
        role = getattr(request.user.userprofile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_DASHBOARD_PAGE_VISIT",
            f"{role.title()} {request.user.username} visited the Dashboard page."
        )

      

        return render(request, 'dashboard.html')

class DashboardDataView(APIView):
    authentication_classes = [DualAuthentication]
    permission_classes = [IsAuthenticated]
    def get(self, request):
        try:
            user = request.user
            user_profile = UserProfile.objects.get(user=user)
            user_role = user_profile.role

            if user_role in ['Manager', 'Team Lead']:
                projects = Project.objects.all()
                tasks = Task.objects.all()
            else:
                projects = Project.objects.filter(project_members__user=user).distinct()
                tasks = Task.objects.filter(Q(assigned_to=user) | Q(created_by=user)).distinct()

            total_tasks = tasks.count()
            completed_tasks = tasks.filter(status='Done').count()
            overdue_tasks = tasks.filter(due_date__lt=timezone.now(), status__in=['To-do', 'In Progress']).count()

            recent_tasks = tasks.order_by('-created_at')[:5].values('title', 'due_date', 'status')

            active_projects_qs = projects.filter(status__in=['Active', 'In Progress'])
            active_projects_count = active_projects_qs.count()
            active_projects_list = active_projects_qs.values('name', 'status', 'progress', 'project_id')

            task_status_counts = tasks.values('status').annotate(count=Count('id'))
            task_status_labels = [item['status'] for item in task_status_counts]
            task_status_data = [item['count'] for item in task_status_counts]

            project_names = [project.name for project in projects]
            project_progress = [project.progress for project in projects]

            data = {
                'total_tasks': total_tasks,
                'completed_tasks': completed_tasks,
                'overdue_tasks': overdue_tasks,
                'active_projects': active_projects_count,
                'recent_tasks': list(recent_tasks),
                'active_projects_list': list(active_projects_list),
                'task_status_labels': task_status_labels,
                'task_status_data': task_status_data,
                'project_names': project_names,
                'project_progress': project_progress
            }

            return Response(data)

        except Exception as e:
            print("Error in DashboardDataView:", str(e))
            return Response({'error': str(e)}, status=500)
        
class NextProjectIdView(APIView):
    def get(self, request):
        try:
            last_project = Project.objects.order_by('-project_id').first()
            if last_project and last_project.project_id:
                try:
                    last_id = int(last_project.project_id)
                    next_id = str(last_id + 1)
                except ValueError:
                    next_id = '1001'
            else:
                next_id = '1001'
            return JsonResponse({'next_project_id': next_id})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
   


def projects_view(request):
    if not request.user.is_authenticated:
        return redirect('login')

    try:
        user_profile = UserProfile.objects.get(user=request.user)
        user_role = user_profile.role
        role_cleaned = user_role.replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_PROJECTS_PAGE_VISIT",
            f"{user_role} {request.user.username} visited the Projects page."
        )
        if user_role in ['Manager', 'Team Lead']:
            projects = Project.objects.select_related('client').all()
        else:
              projects = Project.objects.select_related('client').filter(
        Q(members=request.user) | 
        Q(tasks__assigned_to=request.user)
    ).distinct()

        clients = Client.objects.all()
        return render(request, 'projects.html', {
            'projects': projects,
            'clients': clients,
        })

    except UserProfile.DoesNotExist:
        return redirect('login')
        
class AddProjectView(View):
    authentication_classes = [DualAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        try:
            import re
            data = request.POST
            client_id = data.get('client')
            description = data.get('description', '')
            project_name = data.get('name', '').strip()

            # Validate project name length (min 3, max 40)
            if len(project_name) < 3:
                return JsonResponse({
                    'success': False,
                    'error': 'Project name must be at least 3 characters long.'
                }, status=400)
            
            if len(project_name) > 40:
                return JsonResponse({
                    'success': False,
                    'error': 'Project name cannot exceed 40 characters.'
                }, status=400)

            # Validate project name for leading whitespace
            if project_name and project_name != project_name.lstrip():
                return JsonResponse({
                    'success': False,
                    'error': 'Project name cannot have leading whitespaces.'
                }, status=400)

            # Validate project name allowed characters
            allowed_pattern = re.compile(r'^[a-zA-Z0-9\s\-()]+$')
            if not allowed_pattern.match(project_name):
                return JsonResponse({
                    'success': False,
                    'error': 'Project name can only contain letters, numbers, spaces, hyphens, and parentheses.'
                }, status=400)

            # Check for duplicate project name
            if Project.objects.filter(name__iexact=project_name).exists():
                return JsonResponse({
                    'success': False, 
                    'error': f'A project with the name "{project_name}" already exists. Please choose a different name.'
                }, status=400)

            # Validate description word count (100 words max)
            if description:
                word_count = len(description.strip().split())
                if word_count > 100:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Description cannot exceed 100 words. Current: {word_count} words'
                    }, status=400)

            project = Project.objects.create(
                name=project_name,
                description=description,
                client_id=client_id,
                status=(data.get('status')),
                start_date=data.get('start_date'),
                end_date=data.get('end_date'),
                progress=0,  # Always start at 0
                created_by=request.user
            )
            
            role = getattr(request.user.userprofile, 'role', 'User')
            role_cleaned = role.replace(" ", "_").upper()

            log_type = f"{role_cleaned}_PROJECT_CREATED"
            message = f'{role.title()} {request.user.username} created project "{project.name}".'

            create_activity_log(request.user, log_type, message, project)

            # Call your existing notification function
            create_notification(
                user=request.user,
                title='New Project Created',
                message=f'Project "{project.name}" has been created.',
                notification_type='Project',
                link=f'/projects/{project.id}/view/'
            )

            return JsonResponse({
                'success': True,
                'message': 'Project created successfully',
                'project_id': project.id
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    def get(self, request):
        return JsonResponse({'error': 'Invalid request method'}, status=405)

class ViewProjectView(View):
    def get(self, request, project_id):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        try:
            project = Project.objects.get(id=project_id)
            tasks = Task.objects.filter(project=project).select_related('assigned_to')

            return JsonResponse({
                'project_id': project.project_id,
                'name': project.name,
                'description': project.description,
                'client_name': project.client.name if project.client else 'No Client',
                'status': project.status,
                'start_date': project.start_date.strftime('%Y-%m-%d') if project.start_date else None,
                'end_date': project.end_date.strftime('%Y-%m-%d') if project.end_date else None,
                'progress': project.progress or 0,
                'tasks': [{
                    'title': task.title,
                    'status': task.status,
                    'assigned_to': task.assigned_to.username if task.assigned_to else 'Unassigned'
                } for task in tasks]
            })
        except Project.DoesNotExist:
            return JsonResponse({'error': 'Project not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

class EditProjectView(View):
    def get(self, request, project_id):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        try:
            project = Project.objects.get(id=project_id)
            return JsonResponse({
                'project_id': project.project_id,
                'name': project.name,
                'description': project.description,
                'client_id': project.client.id if project.client else None,
                'status': project.status,
                'start_date': project.start_date.strftime('%Y-%m-%d') if project.start_date else None,
                'end_date': project.end_date.strftime('%Y-%m-%d') if project.end_date else None,
                'progress': project.progress or 0
            })
        except Project.DoesNotExist:
            return JsonResponse({'error': 'Project not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    def post(self, request, project_id):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        try:
            project = Project.objects.get(id=project_id)
            data = request.POST
            description = data.get('description', '')

            # Validate description length (100 words max)
            if description:
                word_count = len(description.strip().split())
                if word_count > 100:
                    return JsonResponse({
                        'success': False, 
                        'error': f'Description cannot exceed 100 words. Current: {word_count} words'
                    }, status=400)

            project.name = data.get('name')
            project.description = description
            project.client_id = data.get('client')
            project.status = data.get('status')
            project.start_date = data.get('start_date') or None
            project.end_date = data.get('end_date') or None
            project.progress = data.get('progress', 0)
            project.save()

            role = getattr(request.user.userprofile, 'role', 'User')
            role_cleaned = role.replace(" ", "_").upper()

            log_type = f"{role_cleaned}_PROJECT_UPDATED"
            message = f'{role.title()} {request.user.username} updated project "{project.name}".'

            create_activity_log(request.user, log_type, message, project)
            return JsonResponse({'success': True, 'message': 'Project updated successfully'})
        except Project.DoesNotExist:
            return JsonResponse({'error': 'Project not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

class DeleteProjectView(View):
    def post(self, request, project_id):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)

        try:
            project = Project.objects.get(id=project_id)
            project_name = project.name
            project.delete()
            role = getattr(request.user.userprofile, 'role', 'User')
            role_cleaned = role.replace(" ", "_").upper()

            log_type = f"{role_cleaned}_PROJECT_DELETED"
            message = f'{role.title()} {request.user.username} deleted project "{project_name}".'

            create_activity_log(request.user, log_type, message)
            return JsonResponse({'success': True, 'message': 'Project deleted successfully'})
        except Project.DoesNotExist:
            return JsonResponse({'error': 'Project not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


class TasksView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')

        ROLE_MAPPING = {
            'Frontend': 'Developer',
            'Backend': 'Developer',
            'AI/ML': 'Developer',
            'Designer (Aximmetry)': 'Designer',
            'Tester': 'Tester',
            'Manager': 'Manager',
        }

        # Log user visit
        try:
            user_profile = UserProfile.objects.get(user=request.user)
            user_role = user_profile.role
            role_cleaned = user_role.replace(" ", "_").upper()
            create_activity_log(
                request.user,
                f"{role_cleaned}_TASKS_PAGE_VISIT",
                f"{user_role} {request.user.username} visited the Tasks page."
            )
        except UserProfile.DoesNotExist:
            return redirect('login')

        # Tasks query
        if user_role in ['Manager', 'Team Lead']:
            tasks = Task.objects.all()
        else:
            tasks = Task.objects.filter(Q(assigned_to=request.user) | Q(created_by=request.user))

        tasks = tasks.select_related('project', 'assigned_to', 'created_by')

        # Filters
        project_filter = request.GET.get('project', '')
        status_filter = request.GET.get('status', '')
        priority_filter = request.GET.get('priority', '')
        assignee_filter = request.GET.get('assignee', '')
        search_term = request.GET.get('search', '')
        sort_value = request.GET.get('sort', 'start_date')
        overdue_flag = request.GET.get('overdue', '')

        if search_term:
            tasks = tasks.filter(
                Q(title__icontains=search_term) |
                Q(description__icontains=search_term) |
                Q(page_name__icontains=search_term) |
                Q(project__name__icontains=search_term) |
                Q(assigned_to__username__icontains=search_term)
            )
        if project_filter:
            tasks = tasks.filter(project_id=project_filter)
        if status_filter:
            tasks = tasks.filter(status=status_filter)
        if priority_filter:
            tasks = tasks.filter(priority=priority_filter)
        if assignee_filter:
            tasks = tasks.filter(assigned_to_id=assignee_filter)

        # Overdue filter (server-side)
        if overdue_flag == '1':
            try:
                today = timezone.now().date()
                tasks = tasks.filter(due_date__lt=today).exclude(status='Done')
            except Exception:
                pass

        # Sorting
        if sort_value == 'priority':
            priority_order = Case(
                When(priority='Urgent', then=Value(1)),
                When(priority='High', then=Value(2)),
                When(priority='Medium', then=Value(3)),
                When(priority='Low', then=Value(4)),
                default=Value(5),
                output_field=IntegerField()
            )
            tasks = tasks.order_by(priority_order)
        else:
            sort_map = {
                'due_date': 'due_date',
                'start_date': 'start_date',
                'created_at': 'created_at',
                'title': 'title'
            }
            tasks = tasks.order_by(sort_map.get(sort_value, 'due_date'))

        # Attach assignee role per task so templates can render "username (Role)"
        try:
            for t in tasks:
                if t.assigned_to_id:
                    # Prefer stored role on task; fallback to membership lookup
                    pm_role = t.assigned_to_role
                    if not pm_role:
                        pm_role = ProjectMember.objects.filter(project_id=t.project_id, user_id=t.assigned_to_id).values_list('role', flat=True).first()
                    t.assignee_role = pm_role
                else:
                    t.assignee_role = None
        except Exception:
            # Fail safe: don't break page rendering if something goes wrong
            pass

        # Manager
        team = Teams.objects.filter(members=request.user).first()
        manager_username = team.team_lead.username if team else None

        # Projects list
        # Limit projects for non-managers to those where the user is associated
        if user_role in ['Manager', 'Team Lead']:
            projects = Project.objects.all()
        else:
            projects = Project.objects.filter(
                Q(teams__members=request.user) |
                Q(project_members__user=request.user) |
                Q(members=request.user)
            ).distinct()

        # Users for chat dropdown
        users = []
        managers = []
        try:
            if user_role == 'Manager':
                users = User.objects.filter(userprofile__role__in=['Developer', 'Tester']).exclude(id=request.user.id)
            else:
                managers = User.objects.filter(userprofile__role='Manager').exclude(id=request.user.id)
        except Exception:
            users = []
            managers = []

        context = {
            'tasks': tasks,
            'projects': projects,
            'users': users,
            'managers': managers,
            'search_term': search_term,
            'project_filter': project_filter,
            'status_filter': status_filter,
            'priority_filter': priority_filter,
            'assignee_filter': assignee_filter,
            'sort_value': sort_value,
            'manager_username': manager_username,
            'today': timezone.now().date(),
        }

        return render(request, 'tasks.html', context)



def assignable_users(request):
    project_id = request.GET.get('project')
    users = []

    if not project_id:
        return JsonResponse({'users': users})

    user_roles_map = defaultdict(list)

    # ---------------- Teams with roles ----------------
    try:
        teams = Teams.objects.filter(project_id=project_id).prefetch_related('members', 'team_role')
        for team in teams:
            role_name = getattr(team.team_role, 'name', None)
            if not role_name:
                continue  # skip if team role is missing

            for member in team.members.all():
                # Add role for this user
                user_roles_map[member.id].append({
                    'user': member,
                    'role': role_name,
                    'source': 'team',
                    'team_id': team.id
                })
    except Exception as e:
        print(f"Error fetching team members: {e}")

    # ---------------- ProjectMember roles ----------------
    try:
        project_members = ProjectMember.objects.filter(project_id=project_id).select_related('user')
        for pm in project_members:
            user = pm.user
            # Roles stored as JSONField or list
            roles_list = pm.roles if isinstance(pm.roles, list) else [pm.roles] if pm.roles else []

            for role in roles_list:
                # Avoid duplicates if already added from Teams
                if role not in [entry['role'] for entry in user_roles_map[user.id]]:
                    user_roles_map[user.id].append({
                        'user': user,
                        'role': role,
                        'source': 'project_member',
                        'team_id': None
                    })
    except Exception as e:
        print(f"Error fetching project members: {e}")

    # ---------------- Build final user list ----------------
    for role_entries in user_roles_map.values():
        for entry in role_entries:
            user = entry['user']
            role = entry['role']
            users.append({
                'id': f"{user.id}:{role}",
                'user_id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'role': role,
                'role_label': f"{user.username} ({role})"
            })

    return JsonResponse({'users': users})


class AddTaskView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.is_authenticated:
            return Response({'success': False, 'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            import re
            data = request.data
            task_title = data.get('title', '').strip()
            project_id = data.get('project')

            # Validate task title length (min 3, max 150)
            if len(task_title) < 3:
                return Response({
                    'success': False,
                    'error': 'Task title must be at least 3 characters long.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if len(task_title) > 150:
                return Response({
                    'success': False,
                    'error': 'Task title cannot exceed 150 characters.'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Validate task title allowed characters
            allowed_pattern = re.compile(r'^[a-zA-Z0-9\s\-().]+$')
            if not allowed_pattern.match(task_title):
                return Response({
                    'success': False,
                    'error': 'Task title can only contain letters, numbers, spaces, hyphens, and parentheses and dots.'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check for duplicate task title in the same project
            if project_id and task_title:
                if Task.objects.filter(project_id=project_id, title__iexact=task_title).exists():
                    return Response({
                        'success': False,
                        'error': f'A task with the title "{task_title}" already exists in this project. Please choose a different title.'
                    }, status=status.HTTP_400_BAD_REQUEST)

            # Extract role if client sent assigned_to as "id:Role"
            raw_assignee = data.get('assigned_to')
            assigned_to_id = None
            assigned_to_role = None

            if raw_assignee and isinstance(raw_assignee, str) and ':' in raw_assignee:
                parts = raw_assignee.split(':')
                assigned_to_id = parts[0]
                assigned_to_role = parts[1]
            else:
                assigned_to_id = raw_assignee

            # Convert empty strings to None for date fields
            start_date = data.get('start_date') or None
            due_date = data.get('due_date') or None

            # Create task
            task = Task.objects.create(
                title=data.get('title'),
                description=data.get('description'),
                page_name=data.get('page_name'),
                project_id=data.get('project'),
                status=data.get('status', 'To-do'),
                priority=data.get('priority', 'Medium'),
                created_by=request.user,
                assigned_to_id=assigned_to_id,
                assigned_to_role=assigned_to_role,
                start_date=start_date,
                due_date=due_date,
                estimated_hours=data.get('estimated_hours') or None
            )

            # Notification for creator
            create_notification(
                user=request.user,
                title='New Task Created',
                message=f'Task "{task.title}" has been created.',
                notification_type='Task',
                link=f'/tasks/{task.id}/view/'
            )

            # Activity log
            role = getattr(request.user.userprofile, 'role', 'User')
            role_cleaned = role.replace(" ", "_").upper()
            log_type = f"{role_cleaned}_TASK_CREATED"
            message = f'{role.title()} {request.user.username} created task "{task.title}".'
            create_activity_log(request.user, log_type, message, task)

            # Notify assignee if assigned
            if assigned_to_id:
                assigned_to = User.objects.get(id=assigned_to_id)
                create_notification(
                    user=assigned_to,
                    title='New Task Assigned',
                    message=f'You have been assigned to task "{task.title}".',
                    notification_type='Task',
                    link=f'/tasks/{task.id}/view/'
                )

            return Response({
                'success': True,
                'message': 'Task created successfully',
                'task_id': task.id
            })

        except Exception as e:
            return Response({'success': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class EditTaskView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, task_id):
        try:
            task = Task.objects.select_related('project', 'assigned_to').get(id=task_id)
            # Find the ProjectMember role for this user in this project
            assigned_to_role = None
            if task.assigned_to:
                # Prefer stored role on task (authoritative)
                assigned_to_role = task.assigned_to_role
                # Do NOT override with membership default if stored role exists
                if not assigned_to_role:
                    pm = ProjectMember.objects.filter(project=task.project, user=task.assigned_to).values_list('role', flat=True).first()
                    assigned_to_role = pm
            data = {
                'title': task.title,
                'project_id': task.project.id,
                'description': task.description,
                'page_name': task.page_name,
                'assigned_to_id': task.assigned_to.id if task.assigned_to else None,
                'assigned_to_role': assigned_to_role,
                'start_date': task.start_date,
                'due_date': task.due_date,
                'status': task.status,
                'priority': task.priority,
                'estimated_hours': task.estimated_hours,
            }
            return Response(data)
        except Task.DoesNotExist:
            return Response({'error': 'Task not found'}, status=status.HTTP_404_NOT_FOUND)
    def post(self, request, task_id):
        try:
            import re
            task = Task.objects.get(id=task_id)
            data = request.data

            # Validate task title if provided
            if 'title' in data:
                task_title = data.get('title', '').strip()
                
                # Validate task title length (min 3, max 150)
                if len(task_title) < 3:
                    return Response({
                        'success': False,
                        'error': 'Task title must be at least 3 characters long.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                if len(task_title) > 150:
                    return Response({
                        'success': False,
                        'error': 'Task title cannot exceed 150 characters.'
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Validate task title allowed characters
                allowed_pattern = re.compile(r'^[a-zA-Z0-9\s\-().]+$')
                if not allowed_pattern.match(task_title):
                    return Response({
                        'success': False,
                        'error': 'Task title can only contain letters, numbers, spaces, hyphens, and parentheses and dots.'
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Check for duplicate task title in the same project
                project_id = data.get('project', task.project_id)
                if project_id and task_title:
                    existing_tasks = Task.objects.filter(project_id=project_id, title__iexact=task_title).exclude(pk=task_id)
                    if existing_tasks.exists():
                        return Response({
                            'success': False,
                            'error': f'A task with the title "{task_title}" already exists in this project. Please choose a different title.'
                        }, status=status.HTTP_400_BAD_REQUEST)

            # Store original assignee for comparison
            original_assignee = task.assigned_to
            original_assignee_role = task.assigned_to_role

            # Update fields if present
            task.title = data.get('title', task.title)
            task.description = data.get('description', task.description)
            task.page_name = data.get('page_name', task.page_name)
            task.project_id = data.get('project', task.project_id)

            # Only update assignee fields if key is provided to avoid unintended unassign on partial updates
            if 'assigned_to' in data:
                assigned_to = data.get('assigned_to')
                # accept values like "userId" or "userId:Role" from client
                if isinstance(assigned_to, str) and ':' in assigned_to:
                    user_id, role = assigned_to.split(':', 1)
                    task.assigned_to_id = user_id
                    task.assigned_to_role = role
                else:
                    # When None/empty is explicitly provided, unassign; otherwise keep as-is
                    if assigned_to:
                        task.assigned_to_id = assigned_to
                    else:
                        task.assigned_to_id = None

            # Only update explicit role if provided; preserves existing role on status-only updates
            if 'assigned_to_role' in data and data.get('assigned_to_role') is not None:
                task.assigned_to_role = data.get('assigned_to_role')

            task.start_date = data.get('start_date') or task.start_date
            task.due_date = data.get('due_date') or task.due_date
            task.status = data.get('status', task.status)
            task.priority = data.get('priority', task.priority)
            task.estimated_hours = data.get('estimated_hours', task.estimated_hours)

            task.save()

            # Check if assignee changed and create automatic comment + notification
            if 'assigned_to' in data:
                new_assignee = task.assigned_to
                if new_assignee != original_assignee:
                    manager_name = request.user.username
                    reassignment_reason = data.get('reassignment_reason', '')

                    if new_assignee:
                        assignee_name = new_assignee.username
                        # Create comment with reason (visible to team leads only)
                        if reassignment_reason:
                            comment_content = f"{manager_name} has reassigned this task to {assignee_name}. Reason: {reassignment_reason}"
                        else:
                            comment_content = f"{manager_name} has reassigned this task to {assignee_name}."
                        # Delete previous reassignment comments for this task
                        # Create the automatic comment
                        Comment.objects.create(
                            user=request.user,
                            content=comment_content,
                            task=task
                        )
                        # Create notification for the new assignee (include reason so UI can display it)
                        notif_message = f"You have been assigned to task: {task.title}"
                        if reassignment_reason:
                            notif_message += f". Reason: {reassignment_reason}"
                        create_notification(
                            user=new_assignee,
                            title="Task Reassigned",
                            message=notif_message,
                            notification_type="Task Assigned",
                            link=f"/tasks/{task.id}/view/"
                        )
                    elif original_assignee:
                        # Task was unassigned
                        comment_content = f"{manager_name} has unassigned this task"
                        Comment.objects.create(
                            user=request.user,
                            content=comment_content,
                            task=task
                        )

            # Log activity
            role = getattr(request.user.userprofile, 'role', 'User')
            role_cleaned = role.replace(" ", "_").upper()

            if 'status' in data and len(data) == 1:
                log_type = f"{role_cleaned}_TASK_STATUS_CHANGED"
                message = f'{role} {request.user.username} changed status of task "{task.title}" to "{task.status}".'
            else:
                log_type = f"{role_cleaned}_TASK_UPDATED"
                message = f'{role} {request.user.username} updated task "{task.title}".'

            create_activity_log(request.user, log_type, message, task)

            return Response({'success': True, 'message': 'Task updated successfully'})

        except Task.DoesNotExist:
            return Response({'error': 'Task not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class ViewTaskView(APIView):
    def get(self, request, task_id):
        if not request.user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            task = Task.objects.select_related('project', 'assigned_to').get(id=task_id)
            created_local = timezone.localtime(task.created_at)  # convert to local timezone

            # Use the stored role from the task, fallback to ProjectMember if not available
            assignee_role = None
            if task.assigned_to:
                # Prefer stored role on task (authoritative)
                assignee_role = task.assigned_to_role
                # Fallback to ProjectMember role if stored role is not available
                if not assignee_role:
                    pm = ProjectMember.objects.filter(project=task.project, user=task.assigned_to).first()
                    if pm:
                        assignee_role = pm.role

            if task.assigned_to:
                if assignee_role:
                    assigned_to_display = f"{task.assigned_to.username} ({assignee_role})"
                else:
                    assigned_to_display = task.assigned_to.username
            else:
                assigned_to_display = 'Unassigned'

            # Filter comments based on user role
            current_user_role = getattr(request.user.userprofile, 'role', 'User')
            is_team_lead = current_user_role in ['Manager', 'Team Lead']
            
            # Get comments and filter visibility
            all_comments = task.comments.all().order_by('created_at')

            # For non-team-lead (Developer/Tester/Designer):
            # - Show all normal comments
            # - Show ONLY the latest reassignment comment
            if not is_team_lead:
                # Prefer the reassignment that assigned the task to the current viewer
                viewer_name = request.user.username
                targeted_reassign = None
                for c in all_comments:
                    text = (c.content or '')
                    if 'has reassigned this task to' in text and f" to {viewer_name}" in text:
                        if targeted_reassign is None or c.created_at > targeted_reassign.created_at:
                            targeted_reassign = c

                if targeted_reassign is None:
                    # Fallback to latest reassignment overall if none specifically to the viewer
                    for c in all_comments:
                        if 'has reassigned this task to' in (c.content or ''):
                            if targeted_reassign is None or c.created_at > targeted_reassign.created_at:
                                targeted_reassign = c

                filtered_comments = []
                for c in all_comments:
                    is_reassign = 'has reassigned this task to' in (c.content or '')
                    if is_reassign and c != targeted_reassign:
                        continue  # show only the one relevant reassignment
                    filtered_comments.append(c)
            else:
                # Managers/Team Leads see everything
                filtered_comments = list(all_comments)

            # Build response with content adjustments
            comments_data = []
            for comment in filtered_comments:
                content_text = comment.content
                if not is_team_lead and 'Reason:' in (content_text or ''):
                    content_text = content_text.split('Reason:')[0].strip()

                comments_data.append({
                    'content': content_text,
                    'user__username': comment.user.username,
                    'created_at': comment.created_at
                })

            data = {
                'title': task.title,
                'project_name': task.project.name,
                'description': task.description,
                'page_name': task.page_name,  # ✅ already included here
                'assigned_to_name': assigned_to_display,
                'start_date': task.start_date.strftime('%Y-%m-%d') if task.start_date else None,
                'due_date': task.due_date.strftime('%Y-%m-%d') if task.due_date else None,
                'status': task.status,
                'priority': task.priority,
                'created_at': created_local.strftime('%d-%m-%Y %I:%M %p'),
                'estimated_hours': task.estimated_hours,
                'comments': comments_data
            }

            # ✅ Force ensure page_name is present and log to confirm
            print("DEBUG → ViewTaskView sending page_name:", task.page_name)
            data["page_name"] = task.page_name

            return Response(data)

        except Task.DoesNotExist:
            return Response({'error': 'Task not found'}, status=status.HTTP_404_NOT_FOUND)


      
class DeleteTaskView(APIView):
    authentication_classes = [SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, task_id):
        if not request.user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            task = Task.objects.get(id=task_id)
            task_title = task.title
            assignee = task.assigned_to
            creator = task.created_by

            # Read reason (but we won't use it in notifications)
            reason = request.POST.get('reason') or request.data.get('reason') or request.data.get('delete_reason') or ''

            # Update existing "New Task Created" notifications to show deletion instead
            existing_notifications = Notification.objects.filter(
                user__in=[creator, assignee] if assignee else [creator],
                title='New Task Created',
                link__icontains=f'/tasks/{task.id}/'
            )

            for notif in existing_notifications:
                notif.title = 'Task Deleted'
                notif.message = f'Task "{task_title}" has been deleted.'
                if reason:
                    notif.message += f' Reason: {reason}'
                notif.read = False  # mark as unread to alert user
                notif.save()

            # Delete the task itself
            task.delete()

            # Activity log
            role = getattr(request.user.userprofile, 'role', 'User')
            role_cleaned = role.replace(" ", "_").upper()
            log_type = f"{role_cleaned}_TASK_DELETED"
            message = f'{role.title()} {request.user.username} deleted task "{task_title}".'
            create_activity_log(request.user, log_type, message)

            return Response({
                'success': True,
                'message': 'Task deleted successfully',
                'reason': reason  # you can still return reason in API if needed
            }, status=status.HTTP_200_OK)

        except Task.DoesNotExist:
            return Response({'error': 'Task not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        
        
class MembersView(View):
    def get(self, request):

       
        if not request.user.is_authenticated:
            return redirect('login')

        role = getattr(request.user.userprofile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_MEMBERS_PAGE_VISIT",
            f"{role.title()} {request.user.username} visited the Members page."
        )

        # Fetch user role
        try:
            user_profile = request.user.userprofile
            user_role = user_profile.role
        except:
            return redirect('login')

        # Fetch projects with teams
        projects_with_teams = Project.objects.filter(teams__isnull=False).distinct()

        # Fetch all relevant teams
        if user_role in ['Manager', 'Team Lead']:
            teams = Teams.objects.filter(project__in=projects_with_teams).select_related('project', 'team_role')
        else:
            teams = Teams.objects.filter(
                members=request.user
            ).select_related('project', 'team_role')

        # Aggregate members per user with per-project roles
        members_list = []
        member_map = defaultdict(lambda: {
            'user': None,
            'projects': {},  # project_id -> { 'project': Project, 'roles': set(), 'joined_at': dt }
        })

        for team in teams:
            for member in team.members.all():
                entry = member_map[member.id]
                entry['user'] = member
                proj_entry = entry['projects'].setdefault(team.project.id, {
                    'project': team.project,
                    'roles': set(),
                    'joined_at': None,
                })

                if team.team_role:
                    proj_entry['roles'].add(team.team_role.name)
                else:
                    profile_role = getattr(member.userprofile, 'role', 'Unknown')
                    if profile_role:
                        proj_entry['roles'].add(profile_role)

                try:
                    pm = ProjectMember.objects.filter(user=member, project=team.project).only('joined_at', 'role').first()
                    if pm:
                        proj_entry['joined_at'] = pm.joined_at
                        if isinstance(pm.role, list):
                            for r in pm.role:
                                if r:
                                    proj_entry['roles'].add(str(r))
                        elif pm.role:
                            proj_entry['roles'].add(str(pm.role))
                except Exception:
                    pass

        for data in member_map.values():
            user = data['user']
            projects_data = []
            role_set_flat = set()
            for pid, info in data['projects'].items():
                roles_str = ", ".join(sorted(info['roles'])) if info['roles'] else 'Member'
                projects_data.append({
                    'id': pid,
                    'name': info['project'].name,
                    'roles': roles_str,
                    'roles_list': sorted(list(info['roles'])),
                    'joined_at': info['joined_at'] or user.date_joined,
                })
                role_set_flat.update(info['roles'])

            members_list.append({
                'id': str(user.id),
                'user': user,
                'projects_data': projects_data,
                'project_names': ", ".join([p['name'] for p in projects_data]),
                'roles': ", ".join(sorted(role_set_flat)) or 'Member',
                'joined_at': user.date_joined,
            })

        form = ProjectMemberForm()

        # Build dynamic role choices from TeamRole and ProjectMember entries
        roles_set = set()
        try:
            team_roles = TeamRole.objects.all().values_list('name', flat=True)
            for r in team_roles:
                if r:
                    roles_set.add(str(r))
        except Exception:
            pass

        try:
            for pm in ProjectMember.objects.all().only('role'):
                role_value = pm.role
                if isinstance(role_value, list):
                    for r in role_value:
                        if r:
                            roles_set.add(str(r))
                elif role_value:
                    roles_set.add(str(role_value))
        except Exception:
            pass

        if not roles_set:
            roles_set.update(['Developer', 'Tester', 'Designer', 'Team Lead', 'Manager'])

        role_choices = sorted(roles_set)

        context = {
            'members': members_list,
            'projects': projects_with_teams,
            'form': form,
            'role_choices': role_choices,
        }
        return render(request, 'members.html', context)
class ViewMember(View):
    def get(self, request, member_id):
        try:
            # member_id may come as "user_id_project_id"
            if '_' in member_id:
                user_id, _ = member_id.split('_', 1)
            else:
                user_id = member_id

            user = User.objects.get(id=user_id)

            # Get all teams where this user is a member
            teams = Teams.objects.filter(members=user).select_related('project', 'team_role')

            projects_data = []
            seen_projects = set()

            for team in teams:
                project = team.project

                # If this project is already processed, append roles to it
                if project.id in seen_projects:
                    for p in projects_data:
                        if p['id'] == project.id:
                            if team.team_role and team.team_role.name not in p['roles_list']:
                                p['roles_list'].append(team.team_role.name)
                            break
                    continue

                # collect all roles for this project
                roles_set = set()
                if team.team_role:
                    roles_set.add(team.team_role.name)
                else:
                    profile_role = getattr(user.userprofile, 'role', 'Member')
                    if profile_role:
                        roles_set.add(profile_role)

                # also include roles from ProjectMember (if available)
                try:
                    pm = ProjectMember.objects.filter(user=user, project=project).only('role', 'joined_at').first()
                    if pm:
                        role_field = pm.role
                        if isinstance(role_field, list):
                            for r in role_field:
                                if r:
                                    roles_set.add(str(r))
                        elif role_field:
                            roles_set.add(str(role_field))
                        joined_at = pm.joined_at
                    else:
                        joined_at = user.date_joined
                except ProjectMember.DoesNotExist:
                    joined_at = user.date_joined

                projects_data.append({
                    'id': project.id,
                    'name': project.name,
                    'roles_list': sorted(list(roles_set)),
                    'joined_at': timezone.localtime(joined_at).strftime('%b %d, %Y, %I:%M %p'),
                })
                seen_projects.add(project.id)

            # fallback join date if no project membership found
            if projects_data:
                first_joined = projects_data[0]['joined_at']
            else:
                first_joined = timezone.localtime(user.date_joined).strftime('%b %d, %Y, %I:%M %p')

            return JsonResponse({
                'user_id': user.id,
                'user_name': user.username,
                'joined_at': first_joined,
                'projects': projects_data
            })

        except User.DoesNotExist:
            return JsonResponse({'error': 'User not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)


def add_team_role(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            role_name = data.get("role_name", "").strip()
            if not role_name:
                return JsonResponse({"success": False, "error": "Role name cannot be empty."})
            
            # Create or get existing role
            role_obj, created = TeamRole.objects.get_or_create(name=role_name)
            return JsonResponse({"success": True, "created": created})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Invalid request."})




class TeamCreateAPIView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')

        role_cleaned = getattr(request.user.userprofile, 'role', 'User').replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_TEAM_PAGE_VISIT",
            f"{request.user.userprofile.role} {request.user.username} visited the Teams page."
        )

        context = {
            'projects': Project.objects.all(),
            'users': User.objects.filter(
                userprofile__role__in=['Developer', 'Tester', 'Team Lead', 'Designer']
            ).distinct(),
            'teams': Teams.objects.all(),
            'roles': TeamRole.objects.all(),
        }
        return render(request, 'teams.html', context)

    def post(self, request):
        # Treat all POST requests as AJAX for this endpoint
        try:
            team_id = request.POST.get('team_id')
            project_id = request.POST.get('project_id')
            team_lead_id = request.POST.get('team_lead')
            team_role_input = request.POST.get('team_role', "").strip()
            member_ids = request.POST.getlist('members')

            # Validation
            if not project_id or not team_lead_id:
                return JsonResponse({"success": False, "error": "Project and Team Lead are required."})

            project = Project.objects.get(id=project_id)
            team_lead = User.objects.get(id=team_lead_id)
            selected_members = User.objects.filter(id__in=member_ids)

            requires_role = selected_members.exclude(userprofile__role='Designer').exists()
            if requires_role and not team_role_input:
                return JsonResponse({"success": False, "error": "Team role is required for non-Designer members."})

            # Get or create TeamRole
            team_role_obj = None
            if team_role_input:
                team_role_obj, _ = TeamRole.objects.get_or_create(name=team_role_input)

            # Enforce DB uniqueness early: same project + same role cannot duplicate regardless of lead/members
            # Only check for duplicates if creating a new team (not updating)
            # Check if team_id is actually provided and not empty
            if not team_id or not str(team_id).strip():
                base_qs = Teams.objects.filter(project=project)
                if team_role_obj:
                    base_qs = base_qs.filter(team_role=team_role_obj)
                else:
                    base_qs = base_qs.filter(team_role__isnull=True)
                if base_qs.exists():
                    return JsonResponse({"success": False, "error": "A team with this project and role already exists."})

            # Duplicate team validation: same project, same lead, same role and same members set
            # Only check for duplicates if creating a new team (not updating)
            if not team_id or not str(team_id).strip():
                def _members_set(team_obj):
                    return set(team_obj.members.values_list('id', flat=True))

                candidate_qs = Teams.objects.filter(project=project, team_lead=team_lead)
                if team_role_obj:
                    candidate_qs = candidate_qs.filter(team_role=team_role_obj)
                else:
                    candidate_qs = candidate_qs.filter(team_role__isnull=True)

                selected_member_ids_set = set(selected_members.values_list('id', flat=True))
                for existing in candidate_qs:
                    if _members_set(existing) == selected_member_ids_set:
                        return JsonResponse({"success": False, "error": "A team with the same project, lead, role and members already exists."})

            # Update or create team
            if team_id and str(team_id).strip():
                team = Teams.objects.get(id=team_id)
                team.project = project
                team.team_lead = team_lead
                team.team_role = team_role_obj if team_role_obj else None
                team.save()
            else:
                team = Teams.objects.create(
                    project=project,
                    team_lead=team_lead,
                    team_role=team_role_obj if team_role_obj else None
                )
            team.members.set(selected_members)

            # Sync ProjectMember roles
            for member in selected_members:
                if team_role_obj:
                    pm, created = ProjectMember.objects.get_or_create(
                        project=project,
                        user=member,
                        defaults={'role': [team_role_obj.name]}
                    )
                    if not created:
                        existing_roles = pm.role or []
                        if team_role_obj.name not in existing_roles:
                            existing_roles.append(team_role_obj.name)
                            pm.role = existing_roles
                            pm.save()
                else:
                    ProjectMember.objects.get_or_create(
                        project=project,
                        user=member,
                        defaults={'role': []}
                    )

            # Log activity
            action = "TEAM_UPDATED" if team_id else "TEAM_CREATED"
            create_activity_log(
                request.user,
                f"{getattr(request.user.userprofile, 'role', 'User').replace(' ', '_').upper()}_{action}",
                f"{request.user.userprofile.role} {request.user.username} {'updated' if team_id else 'created'} team '{team.project.name}'.",
                team
            )

            return JsonResponse({"success": True, "team_id": team.id})

        except Project.DoesNotExist:
            return JsonResponse({"success": False, "error": "Project not found."})
        except User.DoesNotExist:
            return JsonResponse({"success": False, "error": "User not found."})
        except Teams.DoesNotExist:
            return JsonResponse({"success": False, "error": "Team not found."})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

# -----------------------------
# TEAM DELETE VIEW
# -----------------------------
class TeamDeleteView(View):
    def post(self, request, pk):
        try:
            team = get_object_or_404(Teams, id=pk)
            project = team.project

            # Delete associated ProjectMember entries
            for member in team.members.all():
                ProjectMember.objects.filter(user=member, project=project).delete()

            team_name = team.project.name
            team.delete()

            role_cleaned = getattr(request.user.userprofile, 'role', 'User').replace(" ", "_").upper()
            create_activity_log(
                request.user,
                f"{role_cleaned}_TEAM_DELETED",
                f'{request.user.userprofile.role} {request.user.username} deleted team for project "{team_name}".'
            )

            return JsonResponse({"success": True})

        except Teams.DoesNotExist:
            return JsonResponse({"success": False, "error": "Team not found."})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})


class ClientsView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')

        # Determine user role
        try:
            user_role = request.user.userprofile.role
        except UserProfile.DoesNotExist:
            user_role = "User"

        role_cleaned = user_role.replace(" ", "_").upper()
        create_activity_log(
            user=request.user,
            action_type=f"{role_cleaned}_CLIENTS_PAGE_VISIT",
            description=f"{user_role} {request.user.username} visited the Clients page."
        )

        clients = Client.objects.all().order_by('-created_at')
        form = ClientForm()

        formatted_clients = [{
            'id': client.id,
            'name': client.name,
            'email': client.email or 'Not provided',
            'phone': client.phone or 'Not provided',
            'company': client.company or '',
            'address': client.address or 'Not provided',
            'created_at': client.created_at,
        } for client in clients]

        return render(request, 'clients.html', {'clients': formatted_clients, 'form': form})


    def post(self, request):
        if not request.user.is_authenticated:
            return redirect('login')

        action = request.POST.get('action')

        if action == 'add':
            form = ClientForm(request.POST)
            if form.is_valid():
                client = form.save()
                role = getattr(request.user.userprofile, 'role', 'User')
                role_cleaned = role.replace(" ", "_").upper()

                log_type = f"{role_cleaned}_CLIENT_CREATED"
                message = f'{role.title()} {request.user.username} created client "{client.name}".'

                create_activity_log(request.user, log_type, message, client)

                messages.success(request, 'Client added successfully.')
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')

        elif action == 'edit':
            client_id = request.POST.get('client_id')
            try:
                client = Client.objects.get(id=client_id)
                form = ClientForm(request.POST, instance=client)
                if form.is_valid():
                    form.save()
                    role = getattr(request.user.userprofile, 'role', 'User')
                    role_cleaned = role.replace(" ", "_").upper()

                    log_type = f"{role_cleaned}_CLIENT_UPDATED"
                    message = f'{role.title()} {request.user.username} updated client "{client.name}".'

                    create_activity_log(request.user, log_type, message, client)
                    messages.success(request, 'Client updated successfully.')
                else:
                    for field, errors in form.errors.items():
                        for error in errors:
                            messages.error(request, f'{field}: {error}')
            except Client.DoesNotExist:
                messages.error(request, 'Client not found.')

        elif action == 'delete':
            client_id = request.POST.get('client_id')
            try:
                client = Client.objects.get(id=client_id)
                client_name = client.name
                client.delete()
                role = getattr(request.user.userprofile, 'role', 'User')
                role_cleaned = role.replace(" ", "_").upper()

                log_type = f"{role_cleaned}_CLIENT_DELETED"
                message = f'{role.title()} {request.user.username} deleted client "{client_name}".'

                create_activity_log(request.user, log_type, message)
                messages.success(request, 'Client deleted successfully.')
            except Client.DoesNotExist:
                messages.error(request, 'Client not found.')

        return redirect('clients')
class ViewClientAPI(View):
    def get(self, request, client_id):
        if not request.user.is_authenticated:
            return HttpResponseForbidden("Authentication required.")

        try:
            client = Client.objects.get(id=client_id)
            # Convert UTC -> Local timezone (Asia/Kolkata)
            local_time = timezone.localtime(client.created_at)
            formatted_time = local_time.strftime('%b %d, %Y, %I:%M %p')

            return JsonResponse({
                'id': client.id,
                'name': client.name,
                'email': client.email,
                'phone': client.phone,
                'company': client.company,
                'address': client.address,
                'created_at': formatted_time,  # e.g. "Oct 10, 2025, 09:45 PM"
            })
        except Client.DoesNotExist:
            return JsonResponse({'error': 'Client not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

class DocumentView(View):
    def get(self, request, document_id=None):
        if not request.user.is_authenticated:
            return redirect('login')
        
        role = getattr(request.user.userprofile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_DOCUMENTS_PAGE_VISIT",
            f"{role.title()} {request.user.username} visited the Documents page."
        )

        # If document_id is provided, return JSON for that document
        if document_id:
            try:
                doc = Document.objects.select_related('project').get(id=document_id)
                data = {
                    'id': doc.id,
                    'title': doc.title,
                    'project_id': doc.project.id if doc.project else '',
                    'file_name': doc.file.name.split('/')[-1] if doc.file else '',
                    'file_url': doc.file.url if doc.file else '',
                }
                return JsonResponse(data)
            except Document.DoesNotExist:
                return JsonResponse({'error': 'Document not found'}, status=404)

        # Otherwise, render full document list page
        documents = Document.objects.all().select_related('project', 'uploaded_by').order_by('-uploaded_at')
        projects = Project.objects.all()

        formatted_documents = [{
            'id': doc.id,
            'title': doc.title,
            'project': doc.project,
            'file_name': doc.file.name.split('/')[-1] if doc.file else 'No file attached',
            'file_url': doc.file.url if doc.file else None,
            'file_extension': os.path.splitext(doc.file.name)[1].lower() if doc.file else '',
            'uploaded_by': doc.uploaded_by,
            'uploaded_at': timezone.localtime(doc.uploaded_at).strftime('%d/%m/%Y, %I:%M %p') if doc.uploaded_at else 'N/A',
        } for doc in documents]

        return render(request, 'documents.html', {'documents': formatted_documents, 'projects': projects})

    def post(self, request):
        if not request.user.is_authenticated:
            return redirect('login')

        action = request.POST.get('action')

        # Get user role for logging
        role = getattr(request.user.userprofile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()

        if action == 'add':
            try:
                project_id = request.POST.get('project')
                project = Project.objects.get(id=project_id)
                title = (request.POST.get('title') or '').strip()
                # Duplicate title validation (same project + same title, case-insensitive)
                if title:
                    exists = Document.objects.filter(project_id=project_id, title__iexact=title).exists()
                    if exists:
                        messages.error(request, f'A document with the title "{title}" already exists for this project.')
                        return redirect('documents')

                file = request.FILES.get('file')
                if file:
                    document = Document.objects.create(
                        project=project,
                        title=title,
                        file=file,
                        uploaded_by=request.user
                    )

                    log_type = f"{role_cleaned}_DOCUMENT_UPLOADED"
                    message = f'{role.title()} {request.user.username} uploaded document "{document.title}".'
                    create_activity_log(request.user, log_type, message, document)

                    messages.success(request, 'Document uploaded successfully.')
                else:
                    messages.error(request, 'Please select a file to upload.')
            except Exception as e:
                messages.error(request, f'Error uploading document: {str(e)}')

        elif action == 'edit':
            try:
                document = Document.objects.get(id=request.POST.get('document_id'))
                new_title = (request.POST.get('title') or '').strip()
                new_project_id = request.POST.get('project')

                # Duplicate title validation on edit (exclude current document)
                if new_title and new_project_id:
                    exists = Document.objects.filter(
                        project_id=new_project_id,
                        title__iexact=new_title
                    ).exclude(id=document.id).exists()
                    if exists:
                        messages.error(request, f'A document with the title "{new_title}" already exists for this project.')
                        return redirect('documents')

                document.title = new_title
                document.project_id = new_project_id

                if request.FILES.get('file'):
                    if document.file:
                        document.file.delete()
                    document.file = request.FILES.get('file')
                document.save()

                log_type = f"{role_cleaned}_DOCUMENT_UPDATED"
                message = f'{role.title()} {request.user.username} updated document "{document.title}".'
                create_activity_log(request.user, log_type, message, document)

                messages.success(request, 'Document updated successfully.')
            except Document.DoesNotExist:
                messages.error(request, 'Document not found.')
            except Exception as e:
                messages.error(request, f'Error updating document: {str(e)}')

        elif action == 'delete':
            try:
                document = Document.objects.get(id=request.POST.get('document_id'))
                document_title = document.title
                if document.file:
                    document.file.delete()
                document.delete()

                log_type = f"{role_cleaned}_DOCUMENT_DELETED"
                message = f'{role.title()} {request.user.username} deleted document "{document_title}".'
                create_activity_log(request.user, log_type, message)

                messages.success(request, 'Document deleted successfully.')
            except Document.DoesNotExist:
                messages.error(request, 'Document not found.')
            except Exception as e:
                messages.error(request, f'Error deleting document: {str(e)}')

        return redirect('documents')



    
class DownloadDocument(View):
    def get(self, request, document_id):
        if not request.user.is_authenticated:
            return HttpResponseForbidden("You must be logged in to download the file.")

        document = get_object_or_404(Document, id=document_id)
        if document.file:
            try:
                file_path = document.file.path
                with open(file_path, 'rb') as file:
                    response = FileResponse(file)
                    content_type, _ = mimetypes.guess_type(file_path)
                    if content_type:
                        response['Content-Type'] = content_type
                    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
                    return response
            except Exception as e:
                messages.error(request, f'Error downloading file: {str(e)}')
                return redirect('documents')
        raise Http404("File not found")
class PreviewDocument(View):
    @method_decorator(xframe_options_exempt)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get(self, request, document_id):
        if not request.user.is_authenticated:
            return HttpResponse("You must be logged in to preview the file.", status=403)

        document = get_object_or_404(Document, id=document_id)

        if not document.file:
            return HttpResponse("No file attached to this document", status=404)

        file_path = document.file.path

        if not os.path.exists(file_path):
            return HttpResponse("File not found on server", status=404)

        content_type, _ = mimetypes.guess_type(file_path)
        if file_path.lower().endswith(".pdf"):
            content_type = "application/pdf"

        # Open the file without 'with' so FileResponse can stream it
        file = open(file_path, "rb")
        response = FileResponse(file, content_type=content_type)
        response["Content-Disposition"] = f'inline; filename="{os.path.basename(file_path)}"'
        response["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        response["X-Frame-Options"] = "SAMEORIGIN"
        return response

@login_required
def profile_view(request):
    if not request.user.is_authenticated:
        return redirect('login')

    # Log the profile page visit
    try:
        user_role = request.user.userprofile.role
    except UserProfile.DoesNotExist:
        user_role = "User"

    role_cleaned = user_role.replace(" ", "_").upper()
    create_activity_log(
        user=request.user,
        action_type=f"{role_cleaned}_PROFILE_PAGE_VISIT",
        description=f"{user_role} {request.user.username} visited their Profile page."
    )

    profile = request.user.userprofile
    user_teams_with_role = []

    # If Manager, show ALL teams with their leads
    if getattr(profile, 'role', '') == 'Manager':
        all_teams = Teams.objects.all().select_related('project', 'team_lead', 'team_role')
        for team in all_teams:
            user_teams_with_role.append({
                'id': team.id,
                'label': f"{team.project.name} ({team.team_role.name if team.team_role else 'Member'}) Lead: {team.team_lead.username}"
            })
        context = {
            'user': request.user,
            'profile': profile,
            'teams': Teams.objects.all(),
            'user_teams_with_role': user_teams_with_role,
        }
        return render(request, 'profile.html', context)

    # ✅ Teams where user is a member
    member_teams = Teams.objects.filter(members=request.user).select_related('project', 'team_lead')
    for team in member_teams:
        # Get ALL roles the user has in this project
        pms = ProjectMember.objects.filter(project=team.project, user=request.user)
        if pms.exists():
            for pm in pms:
                user_teams_with_role.append({
                    'id': team.id,
                    'label': f"{team.project.name} ({pm.role}) - Manager: {team.team_lead.username}"
                })
        else:
            # fallback if no ProjectMember role is defined
            user_teams_with_role.append({
                'id': team.id,
                'label': f"{team.project.name} (Member) - Manager: {team.team_lead.username}"
            })

    # ✅ Teams where user is manager
    managed_teams = Teams.objects.filter(team_lead=request.user).select_related('project')
    for team in managed_teams:
        if not any(t['id'] == team.id for t in user_teams_with_role):
            user_teams_with_role.append({
                'id': team.id,
                'label': f"{team.project.name} (Manager)"
            })

    context = {
        'user': request.user,
        'profile': profile,
        'teams': Teams.objects.all(),
        'user_teams_with_role': user_teams_with_role,
    }
    return render(request, 'profile.html', context)




class NotificationListView(APIView):
    authentication_classes = [SessionAuthentication, BasicAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Ensure user is authenticated
        if not request.user.is_authenticated:
            return redirect('login')

        # Get user role
        try:
            user_role = request.user.userprofile.role
        except UserProfile.DoesNotExist:
            user_role = "User"

        role_cleaned = user_role.replace(" ", "_").upper()

        # Detect AJAX or background fetch request
        accept_header = request.headers.get('Accept', '')
        sec_fetch_mode = request.META.get('HTTP_SEC_FETCH_MODE', '')
        is_ajax = (
            request.headers.get('x-requested-with') == 'XMLHttpRequest'
            or 'application/json' in accept_header
            or sec_fetch_mode in ('cors', 'no-cors')
        )

        # Only log if it's a real page visit (not an AJAX call)
        if not is_ajax:
            create_activity_log(
                user=request.user,
                action_type=f"{role_cleaned}_NOTIFICATIONS_PAGE_VISIT",
                description=f"{user_role} {request.user.username} visited the Notifications page."
            )

        # Fetch notifications for the logged-in user
        notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
        unread_count = notifications.filter(read=False).count()

        # If it's an AJAX/fetch request, return JSON instead of rendering a page
        if is_ajax:
            data = [
                {
                    'id': n.id,
                    'title': n.title,
                    'message': n.message,
                    'notification_type': n.notification_type,
                    'link': n.link,
                    'read': n.read,
                    'created_at': n.created_at.strftime('%Y-%m-%d %H:%M'),
                }
                for n in notifications
            ]
            return JsonResponse(data, safe=False)

        # Render notifications page for normal visits
        context = {
            'notifications': notifications,
            'unread_count': unread_count,
        }
        return render(request, 'notifications.html', context)
 
class ClearSelectedNotificationsView(View):
    def post(self, request):
        ids = request.POST.getlist('ids[]', [])
        if not ids:
            return JsonResponse({'error': 'No notifications selected'}, status=400)

        # Delete only current user's notifications
        deleted_count, _ = Notification.objects.filter(id__in=ids, user=request.user).delete()

        if deleted_count > 0:
            return JsonResponse({'success': True, 'message': 'Selected notifications cleared successfully'})
        else:
            return JsonResponse({'error': 'No matching notifications found'}, status=404)


class NotificationCountAPIView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, BasicAuthentication]

    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'count': 0})
        count = Notification.objects.filter(user=request.user, read=False).count()
        return JsonResponse({'count': count})

class TaskNotificationCountAPIView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, BasicAuthentication]

    def get(self, request, task_id):
        if not request.user.is_authenticated:
            return JsonResponse({'count': 0})
        # Get notifications related to this specific task
        task_notifications = Notification.objects.filter(
            user=request.user, 
            read=False,
            link__contains=f"/tasks/{task_id}/"
        ).count()
        return JsonResponse({'count': task_notifications})

class MarkNotificationReadView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [SessionAuthentication, BasicAuthentication]

    def post(self, request, notification_id):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Authentication required'}, status=401)
        
        try:
            notification = Notification.objects.get(id=notification_id, user=request.user)
            # Only flip and broadcast if it was unread
            if not notification.read:
                notification.read = True
                notification.save()

                # Broadcast updated unread count to this user's notification channel group
                try:
                    from channels.layers import get_channel_layer
                    from asgiref.sync import async_to_sync

                    channel_layer = get_channel_layer()
                    if channel_layer:
                        count = Notification.objects.filter(user=request.user, read=False).count()
                        async_to_sync(channel_layer.group_send)(
                            f'notifications_{request.user.id}',
                            {
                                'type': 'notification_update',
                                'count': count,
                                'message': 'Notification marked as read'
                            }
                        )
                except Exception as e:
                    # Do not fail the request if websocket broadcast fails
                    print(f"WebSocket broadcast failed on mark-read: {e}")

            return JsonResponse({'success': True, 'message': 'Notification marked as read'})
        except Notification.DoesNotExist:
            return JsonResponse({'error': 'Notification not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
class ViewUserView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            profile = getattr(user, "userprofile", None)

            # Prepare response data
            data = {
                "username": getattr(user, "username", "-"),
                "email": getattr(profile, "email", user.email or "-") if profile else user.email or "-",
                "phone": getattr(profile, "phone", "-") if profile else "-",
                "role": getattr(profile, "role", "-") if profile else "-",
                "empid": getattr(profile, "empid", "-") if profile else "-",
                "is_superuser": user.is_superuser,
                "date_joined": user.date_joined.strftime("%Y-%m-%d %H:%M") if user.date_joined else "-",
                "profile_picture_url": profile.profile_picture.url if profile and getattr(profile, "profile_picture", None) else None,
                "teams": [
                    getattr(team.project, "name", "-")
                    for team in getattr(user, "teams_set", Teams.objects.none()).all()
                ],
            }

            return Response(data)
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


def create_notification(user, title, message, notification_type, link=None):
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync
    
    print(f"Creating notification for user {user.username}: {title}")
    
    # Create the notification
    notification = Notification.objects.create(
        user=user,
        title=title,
        message=message,
        notification_type=notification_type,
        link=link
    )
    
    print(f"Notification created with ID: {notification.id}")
    
    # Send real-time update to the user
    try:
        channel_layer = get_channel_layer()
        print(f"Channel layer: {channel_layer}")
        if channel_layer:
            # Get updated count
            count = Notification.objects.filter(user=user, read=False).count()
            print(f"Updated notification count for user {user.username}: {count}")
            
            # Send update to user's notification group
            group_name = f'notifications_{user.id}'
            print(f"Sending update to group: {group_name}")
            async_to_sync(channel_layer.group_send)(
                group_name,
                {
                    'type': 'notification_update',
                    'count': count
                }
            )
            print("WebSocket update sent successfully")
    except Exception as e:
        # If WebSocket fails, don't break the notification creation
        print(f"WebSocket notification update failed: {e}")
    
    return notification



class TaskListView(APIView):
    def get(self, request, *args, **kwargs):
        # Get all tasks
        tasks = Task.objects.all()
        serializer = TaskSerializer(tasks, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TaskDistributionAPIView(APIView):
    def get(self, request, *args, **kwargs):
        try:
            # Get all projects
            projects = Project.objects.all()
            project_data = []
            
            for project in projects:
                # Get all team members for this project
                project_members = ProjectMember.objects.filter(project=project)
                
                assignees_data = []
                for member in project_members:
                    user = member.user
                    
                    # Count total tasks assigned to this user in this project
                    total_tasks = Task.objects.filter(
                        project=project,
                        assigned_to=user
                    ).count()
                    
                    # Count completed tasks assigned to this user in this project
                    completed_tasks = Task.objects.filter(
                        project=project,
                        assigned_to=user,
                        status='Done'
                    ).count()
                    
                    assignees_data.append({
                        'name': user.get_full_name() or user.username,
                        'total_tasks': total_tasks,
                        'completed_tasks': completed_tasks
                    })
                
                # Only include projects that have members with tasks
                if assignees_data and any(assignee['total_tasks'] > 0 for assignee in assignees_data):
                    project_data.append({
                        'id': project.id,
                        'name': project.name,
                        'assignees': assignees_data
                    })
            
            return Response({
                'success': True,
                'projects': project_data
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)







class SettingsView(View):

    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        role = getattr(request.user.userprofile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_SETTINGS_PAGE_VISIT",
            f"{role.title()} {request.user.username} visited the Settings page."
        )

        profile = request.user.userprofile

        # Build dropdown options based on role
        if request.user.is_superuser:
            user_dropdown_options = []

        elif profile.role == 'Manager':
            # Managers should see all teams with their leads
            teams_qs = Teams.objects.all().select_related('project', 'team_role', 'team_lead')
            user_dropdown_options = [
                {
                    'id': team.id,
                    'label': f"{team.project.name} ({team.team_role.name if team.team_role else 'Member'}) - Lead: {team.team_lead.username}"
                }
                for team in teams_qs
            ]

        elif profile.role == 'Team Lead':
            # Team Lead: teams they lead OR where they are members
            teams_qs = Teams.objects.filter(Q(team_lead=request.user) | Q(members=request.user)).select_related('project', 'team_lead', 'team_role').distinct()
            user_dropdown_options = []
            for team in teams_qs:
                pms = ProjectMember.objects.filter(project=team.project, user=request.user)
                if pms.exists():
                    for pm in pms:
                        user_dropdown_options.append({
                            'id': team.id,
                            'label': f"{team.project.name} ({pm.role}) - Lead: {team.team_lead.username}"
                        })
                else:
                    user_dropdown_options.append({
                        'id': team.id,
                        'label': f"{team.project.name} ({team.team_role.name if team.team_role else 'Member'}) - Lead: {team.team_lead.username}"
                    })

        else:
            # Developer/Tester/Designer: teams where they are members
            user_dropdown_options = []
            member_teams = Teams.objects.filter(members=request.user).select_related('project', 'team_lead', 'team_role')
            for team in member_teams:
                # Fetch ALL roles of this user in this project
                pms = ProjectMember.objects.filter(project=team.project, user=request.user)
                if pms.exists():
                    for pm in pms:
                        user_dropdown_options.append({
                            'id': team.id,
                            'label': f"{team.project.name} ({pm.role}) - Lead: {team.team_lead.username}"
                        })
                else:
                    # fallback if no role entry found
                    user_dropdown_options.append({
                        'id': team.id,
                        'label': f"{team.project.name} (Member) - Lead: {team.team_lead.username}"
                    })

        # Role memberships for badges
        user_role_memberships = [
            {'project_name': pm.project.name, 'label': f"{pm.project.name} ({pm.role})"}
            for pm in ProjectMember.objects.filter(user=request.user).select_related('project')
        ]

        context = {
            'user': request.user,
            'profile': profile,
            'is_admin': request.user.is_superuser,
            'users': User.objects.all() if request.user.is_superuser else None,
            'user_dropdown_options': user_dropdown_options,
            'modules': [
                {'id': 'documents', 'name': 'Documents'},
                {'id': 'clients', 'name': 'Clients'},
                {'id': 'reports', 'name': 'Reports'},
                {'id': 'teams', 'name': 'Teams'},
            ],
            'user_role_memberships': user_role_memberships,
        }
        return render(request, 'settings.html', context)


    def post(self, request):
        action = request.POST.get('action')
        user = request.user
        profile = user.userprofile
        role = getattr(profile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()

        if action == 'account':
            try:
                old_username = user.username
                old_email = profile.email
                old_phone = profile.phone
                old_role = profile.role

                # Update User fields
                user.username = request.POST.get('username')
                user.save()
                if user.username != old_username:
                    create_activity_log(user, 'ACCOUNT_USERNAME_UPDATED',
                                        f'User changed username from {old_username} to {user.username}.')

                # Update UserProfile fields
                profile.email = request.POST.get('email')
                profile.phone = request.POST.get('phone')
                # Only update role if provided (select may be disabled for non-Managers)
                if request.POST.get('role') is not None and request.POST.get('role') != '':
                    profile.role = request.POST.get('role')

                # Handle profile picture
                if request.FILES.get('profile_picture'):
                    if profile.profile_picture and os.path.exists(profile.profile_picture.path):
                        default_storage.delete(profile.profile_picture.path)
                    profile.profile_picture = request.FILES['profile_picture']
                profile.save()

                # Log field changes
                if profile.email != old_email:
                    create_activity_log(user, f'{role_cleaned}_ACCOUNT_EMAIL_UPDATED',
                                        f'{role.title()} changed email from {old_email} to {profile.email}.')
                if profile.phone != old_phone:
                    create_activity_log(user, f'{role_cleaned}_ACCOUNT_PHONE_UPDATED',
                                        f'{role.title()} changed phone from {old_phone} to {profile.phone}.')
                if profile.role != old_role:
                    create_activity_log(user, f'{role_cleaned}_ACCOUNT_ROLE_UPDATED',
                                        f'{role.title()} changed role from {old_role} to {profile.role}.')

                # --------------------------
                # Handle team assignment ONLY if user has no team
                # --------------------------
                team_id = request.POST.get('team')
                if team_id:
                    try:
                        new_team = Teams.objects.get(id=team_id)
                        # Only assign the user if they are not already part of this team
                        if not new_team.members.filter(id=user.id).exists():
                            new_team.members.add(user)
                            new_team.save()
                            create_activity_log(user, f'{role_cleaned}_ACCOUNT_TEAM_ASSIGNED',
                                                f'{role.title()} was assigned to team {new_team.project.name}.')
                    except Teams.DoesNotExist:
                        pass


                create_activity_log(user, f'{role_cleaned}_ACCOUNT_SETTINGS_UPDATED',
                                    f'{role.title()} {user.username} updated their account settings.')

                messages.success(request, 'Account settings updated successfully.')

            except Exception as e:
                messages.error(request, f'Error updating account settings: {str(e)}')

        elif action == 'appearance':
            try:
                old_theme = profile.theme
                old_primary_color = profile.primary_color
                old_font_size = profile.font_size

                theme = request.POST.get('theme')
                if theme in ['light', 'dark']:
                    profile.theme = theme
                profile.primary_color = request.POST.get('primary_color', '#1e88e5')
                font_size = request.POST.get('font_size')
                if font_size in ['small', 'medium', 'large']:
                    profile.font_size = font_size
                profile.save()

                if profile.theme != old_theme:
                    create_activity_log(user, f'{role_cleaned}_APPEARANCE_THEME_UPDATED',
                                        f'{role.title()} changed theme from {old_theme} to {profile.theme}.')
                if profile.primary_color != old_primary_color:
                    create_activity_log(user, f'{role_cleaned}_APPEARANCE_COLOR_UPDATED',
                                        f'{role.title()} changed primary color from {old_primary_color} to {profile.primary_color}.')
                if profile.font_size != old_font_size:
                    create_activity_log(user, f'{role_cleaned}_APPEARANCE_FONT_SIZE_UPDATED',
                                        f'{role.title()} changed font size from {old_font_size} to {profile.font_size}.')

                create_activity_log(user, f'{role_cleaned}_APPEARANCE_SETTINGS_UPDATED',
                                    f'{role.title()} {user.username} updated their appearance settings.')

                messages.success(request, 'Appearance settings updated successfully.')

            except Exception as e:
                messages.error(request, f'Error updating appearance settings: {str(e)}')

        return redirect('settings')


# New: Add existing employee to another team (no re-registration)
class AddEmployeeToTeamView(View):
    def post(self, request):
        if not request.user.is_authenticated:
            return HttpResponseForbidden("You must be logged in to perform this action.")

        try:
            empid = request.POST.get('empid')
            team_id = request.POST.get('team_id')
            make_primary = request.POST.get('make_primary') == 'true'

            if not empid or not team_id:
                return JsonResponse({'success': False, 'error': 'empid and team_id are required'}, status=400)

            profile = UserProfile.objects.get(empid=empid)
            team = Teams.objects.get(id=team_id)

            if make_primary:
                profile.team = team
                profile.save()
            else:
                profile.additional_teams.add(team)

            role = getattr(request.user.userprofile, 'role', 'User')
            role_cleaned = role.replace(" ", "_").upper()
            action = 'PRIMARY_TEAM_SET' if make_primary else 'ADDED_TO_ADDITIONAL_TEAM'
            create_activity_log(
                request.user,
                f"{role_cleaned}_{action}",
                f"{role.title()} {request.user.username} added {profile.user.username} to team '{team.__str__()}'.",
                team
            )

            return JsonResponse({'success': True})
        except UserProfile.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Employee not found'}, status=404)
        except Teams.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Team not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

def forgot_password(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')

            if not email:
                return JsonResponse({'error': 'Email is required'}, status=400)

            user = User.objects.filter(email=email).first()
            if not user:
                return JsonResponse({'error': 'No user found with this email'}, status=404)

            # Create or retrieve auth token
            token, _ = Token.objects.get_or_create(user=user)

            # Get current domain dynamically
            current_site = get_current_site(request)
            domain = current_site.domain
            protocol = 'https' if request.is_secure() else 'http'
            
            # Build reset link
            reset_path = reverse('reset_password') + f"?token={token.key}"
            reset_link = f"{protocol}://{domain}{reset_path}"

            # Send the email
            send_mail(
                subject='Reset Your Password',
                message=f'Click the link below to reset your password:\n\n{reset_link}',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                html_message=f'''
                    <p>Hello {user.username},</p>
                    <p>Click the link below to reset your password:</p>
                    <p><a href="{reset_link}">{reset_link}</a></p>
                    <p>If you didn't request this, please ignore this email.</p>
                ''',
                fail_silently=False,
            )

            return JsonResponse({'message': 'Password reset link sent!'})

        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

    return render(request, 'forgot_password.html')
def reset_password(request):
    token_key = request.GET.get('token') if request.method == 'GET' else request.POST.get('token')

    if request.method == 'GET':
        return render(request, 'reset_password.html', {'token': token_key})

    if request.method == 'POST':
        password = request.POST.get('password')

        if not password or not token_key:
            return HttpResponse("Invalid request", status=400)

        token = Token.objects.filter(key=token_key).first()
        if not token:
            return HttpResponse("Invalid or expired token", status=404)

        user = token.user
        print("user",user)
        user.password = make_password(password)
        print('password',password)
        user.save()

        register_user = UserProfile.objects.get(user=user)
        register_user.password = make_password(password)
        print('password',password)
        register_user.save()

        create_activity_log(user, 'PASSWORD_RESET', f'User {user.username} reset their password.')

        # Invalidate token after password reset (optional security step)
        token.delete()

        # return HttpResponse("Password reset successful. You can now log in.")
        return render(request, 'login.html')




class ReportsView(View):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')

        
        url_name = request.resolver_match.url_name
        print("Resolved URL:", url_name)
        role = getattr(request.user.userprofile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_REPORTS_PAGE_VISIT",
            f"{role.title()} {request.user.username} visited the Reports page."
        )


        if url_name == 'reports':
            return render(request, 'reports.html')
        elif url_name == 'project_reports':
            return self.get_project_reports(request)
        elif url_name == 'task_reports':
            return self.get_task_reports(request)
        elif url_name == 'team_reports':
            return self.get_team_reports(request)
        elif url_name == 'time_reports':
            return self.get_time_reports(request)
        elif url_name == 'activity_reports':
            return self.get_activity_reports(request)
        elif url_name == 'burndown_reports':
            return self.get_burndown_reports(request)
        elif url_name == 'export_report':
            report_type = kwargs.get('report_type')
            format = kwargs.get('format')
            return self.export_report(request, report_type, format)
        else:
            return JsonResponse({'error': 'Invalid report type'}, status=400)
    def get_project_reports(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        projects = Project.objects.all()
        if start_date and end_date:
            projects = projects.filter(
                created_at__date__range=[start_date, end_date]
            )
        
        data = {
            'projects': [{
                'name': p.name,
                'status': p.status,
                'priority': p.priority,
                'client': p.client.name if p.client else 'No Client',
                'start_date': p.start_date.strftime('%Y-%m-%d') if p.start_date else None,
                'end_date': p.end_date.strftime('%Y-%m-%d') if p.end_date else None,
                'deadline': p.deadline.strftime('%Y-%m-%d') if p.deadline else None,
                'progress': p.progress,
                'budget': float(p.budget) if p.budget else None,
                'total_tasks': p.tasks.count(),
                'completed_tasks': p.tasks.filter(status='Done').count(),
                'team_members': [{
                    'name': member.user.username,
                    'role': member.role
                } for member in p.project_members.all()],
                'documents_count': p.documents.count()
            } for p in projects],
            'summary': {
                'total_projects': projects.count(),
                'projects_by_status': dict(projects.values_list('status').annotate(count=models.Count('id'))),
                'projects_by_priority': dict(projects.values_list('priority').annotate(count=models.Count('id'))),
                'average_progress': projects.aggregate(avg_progress=models.Avg('progress'))['avg_progress'] or 0
            }
        }
        return JsonResponse(data)

    def get_task_reports(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        tasks = Task.objects.all()
        if start_date and end_date:
            tasks = tasks.filter(
                created_at__date__range=[start_date, end_date]
            )
        
        status_counts = {
            'todo': tasks.filter(status='To-do').count(),
            'in_progress': tasks.filter(status='In Progress').count(),
            'review': tasks.filter(status='Review').count(),
            'done': tasks.filter(status='Done').count()
        }
        
        priority_counts = {
            'urgent': tasks.filter(priority='Urgent').count(),
            'high': tasks.filter(priority='High').count(),
            'medium': tasks.filter(priority='Medium').count(),
            'low': tasks.filter(priority='Low').count()
        }
        
        overdue_tasks = tasks.filter(
            due_date__lt=timezone.now(),
            status__in=['To-do', 'In Progress']
        ).values('title', 'due_date', 'status', 'priority', 'project__name', 'assigned_to__username')
        
        data = {
            'status_counts': status_counts,
            'priority_counts': priority_counts,
            'overdue_tasks': list(overdue_tasks),
            'summary': {
                'total_tasks': tasks.count(),
                'completed_tasks': tasks.filter(status='Done').count(),
                'tasks_by_project': dict(tasks.values_list('project__name').annotate(count=models.Count('id'))),
                'average_completion_time': tasks.filter(status='Done').aggregate(
                    avg_time=models.Avg(models.F('updated_at') - models.F('created_at'))
                )['avg_time'],
                'tasks_by_assignee': dict(tasks.values_list('assigned_to__username').annotate(count=models.Count('id')))
            }
        }
        return JsonResponse(data)

    def get_team_reports(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        period = request.GET.get('period', 'monthly')  # weekly | monthly | yearly

        # Default date range: last 30 days
        if not start_date or not end_date:
            end = timezone.now().date()
            start = end - timezone.timedelta(days=30)
            start_date = start.strftime('%Y-%m-%d')
            end_date = end.strftime('%Y-%m-%d')

        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Build buckets for the Y axis based on period
        buckets = []
        bucket_label = 'Month'
        current = start_dt
        if period == 'weekly':
            bucket_label = 'Week'
            while current <= end_dt:
                week_start = current - timezone.timedelta(days=current.weekday())
                week_end = week_start + timezone.timedelta(days=6)
                buckets.append((max(week_start, start_dt), min(week_end, end_dt)))
                current = week_end + timezone.timedelta(days=1)
        elif period == 'yearly':
            bucket_label = 'Year'
            year = start_dt.year
            while year <= end_dt.year:
                y_start = datetime(year, 1, 1).date()
                y_end = datetime(year, 12, 31).date()
                buckets.append((max(y_start, start_dt), min(y_end, end_dt)))
                year += 1
        else:
            # monthly (default)
            while current <= end_dt:
                m_start = current.replace(day=1)
                if m_start.month == 12:
                    next_month_start = datetime(m_start.year + 1, 1, 1).date()
                else:
                    next_month_start = datetime(m_start.year, m_start.month + 1, 1).date()
                m_end = next_month_start - timezone.timedelta(days=1)
                buckets.append((max(m_start, start_dt), min(m_end, end_dt)))
                current = next_month_start

        # Collect members (users) under all teams or with assigned tasks
        team_members = User.objects.filter(
            Q(team_members__isnull=False) | Q(assigned_tasks__isnull=False)
        ).distinct()
        # Fallback: users who have tasks in the period
        if not team_members.exists():
            team_members = User.objects.filter(
                assigned_tasks__created_at__date__gte=start_dt,
                assigned_tasks__created_at__date__lte=end_dt
            ).distinct()

        members_summary = []
        scatter = []
        palette = [
            'rgba(244,67,54,0.6)', 'rgba(33,150,243,0.6)', 'rgba(76,175,80,0.6)',
            'rgba(255,152,0,0.6)', 'rgba(156,39,176,0.6)', 'rgba(0,188,212,0.6)'
        ]

        for idx, member in enumerate(team_members):
            # Tasks within range for the user
            user_tasks = Task.objects.filter(
                assigned_to=member,
                created_at__date__gte=start_dt,
                created_at__date__lte=end_dt
            )
            completed = user_tasks.filter(status='Done').count()
            pending = user_tasks.exclude(status='Done').count()
            members_summary.append({
                'name': member.username,
                'completed': completed,
                'pending': pending
            })

            # Build scatter points across buckets
            points = []
            for b_idx, (b_start, b_end) in enumerate(buckets, start=1):
                bucket_tasks = user_tasks.filter(created_at__date__gte=b_start, created_at__date__lte=b_end)
                total_tasks = bucket_tasks.count()
                done = bucket_tasks.filter(status='Done').count()
                completion_rate = (done / total_tasks * 100) if total_tasks > 0 else 0
                points.append({
                    'total_tasks': total_tasks,
                    'bucket_index': b_idx,
                    'completion_rate': completion_rate
                })

            scatter.append({
                'member': member.username,
                'color': palette[idx % len(palette)],
                'points': points
            })

        # Resource utilization metrics per member (without TimeEntry)
        utilization = []
        for member in team_members:
            # Assigned hours from tasks' estimated_hours
            assigned_hours = Task.objects.filter(
                assigned_to=member,
                created_at__date__gte=start_dt,
                created_at__date__lte=end_dt
            ).aggregate(total=models.Sum('estimated_hours'))['total'] or 0

            utilization_pct = 0.0  # Without TimeEntry, cannot calculate actual utilization
            status_flag = 'balanced'

            utilization.append({
                'name': member.username,
                'assigned_hours': float(assigned_hours or 0),
                'logged_hours': None,  # No logged hours
                'utilization_pct': utilization_pct,
                'status': status_flag
            })

        # Manager-focused team performance line series over buckets
        # Compute per bucket team completion rate and active projects count
        line_labels = []
        completion_series = []
        projects_series = []
        for (b_start, b_end) in buckets:
            line_labels.append(f"{b_start.strftime('%Y-%m-%d')}\n{b_end.strftime('%Y-%m-%d')}")
            bucket_tasks_all = Task.objects.filter(
                created_at__date__gte=b_start,
                created_at__date__lte=b_end
            )
            total_bucket = bucket_tasks_all.count()
            done_bucket = bucket_tasks_all.filter(status='Done').count()
            completion_series.append(round(((done_bucket / total_bucket) * 100) if total_bucket else 0, 2))
            projects_series.append(
                Project.objects.filter(
                    tasks__created_at__date__gte=b_start,
                    tasks__created_at__date__lte=b_end
                ).distinct().count()
            )

        data = {
            'members': members_summary,
            'scatter': scatter,
            'bucket_label': bucket_label,
            'utilization': utilization,
            'manager_line': {
                'labels': line_labels,
                'completion_rate': completion_series,
                'active_projects': projects_series
            }
        }
        return JsonResponse(data)


    def get_burndown_reports(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')

        # Default last 30 days
        if not start_date or not end_date:
            end = timezone.now().date()
            start = end - timezone.timedelta(days=30)
            start_date = start.strftime('%Y-%m-%d')
            end_date = end.strftime('%Y-%m-%d')

        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Build daily series
        series = []
        day = start_dt
        total_tasks_all = Task.objects.filter(created_at__date__lte=end_dt).count()
        completed_cumulative = 0
        while day <= end_dt:
            created_to_day = Task.objects.filter(created_at__date__lte=day).count()
            done_to_day = Task.objects.filter(status='Done', updated_at__date__lte=day).count()
            remaining = max(created_to_day - done_to_day, 0)
            completed_cumulative = done_to_day
            series.append({
                'date': day.strftime('%Y-%m-%d'),
                'remaining': remaining,
                'completed': completed_cumulative
            })
            day += timezone.timedelta(days=1)

        return JsonResponse({'series': series, 'total_tasks': total_tasks_all})

    def get_activity_reports(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        notifications = Notification.objects.all()
        comments = Comment.objects.all()
        
        if start_date and end_date:
            notifications = notifications.filter(
                created_at__date__range=[start_date, end_date]
            )
            comments = comments.filter(
                created_at__date__range=[start_date, end_date]
            )
        
        data = {
            'notifications': [{
                'title': n.title,
                'message': n.message,
                'type': n.notification_type,
                'user': n.user.username,
                'timestamp': n.created_at.strftime('%Y-%m-%d %H:%M'),
                'read': n.read
            } for n in notifications],
            'comments': [{
                'content': c.content,
                'user': c.user.username,
                'project': c.project.name if c.project else None,
                'task': c.task.title if c.task else None,
                'timestamp': c.created_at.strftime('%Y-%m-%d %H:%M')
            } for c in comments],
            'summary': {
                'total_notifications': notifications.count(),
                'unread_notifications': notifications.filter(read=False).count(),
                'notifications_by_type': dict(notifications.values_list('notification_type').annotate(count=models.Count('id'))),
                'total_comments': comments.count(),
                'comments_by_user': dict(comments.values_list('user__username').annotate(count=models.Count('id')))
            }
        }
        return JsonResponse(data)

    def export_report(self, request, report_type, format):
        # Get the data for the report
        if report_type == 'project':
            data = self.get_project_reports(request)
        elif report_type == 'task':
            data = self.get_task_reports(request)
        elif report_type == 'team':
            data = self.get_team_reports(request)
        elif report_type == 'time':
            data = self.get_time_reports(request)
        elif report_type == 'activity':
            data = self.get_activity_reports(request)
        elif report_type == 'burndown':
            data = self.get_burndown_reports(request)
        else:
            return JsonResponse({'error': 'Invalid report type'}, status=400)

        # If the data is a JsonResponse, extract the data dictionary
        if isinstance(data, JsonResponse):
            import json
            data = json.loads(data.content)

        if format == 'pdf':
            return self.generate_pdf(report_type, data)
        elif format == 'excel':
            return self.generate_excel(report_type, data)
        else:
            return JsonResponse({'error': 'Invalid format'}, status=400)

    def generate_pdf(self, report_type, data):
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, f"{report_type.capitalize()} Report")
        p.setFont("Helvetica", 10)
        y = height - 80

        if report_type == 'project':
            p.drawString(50, y, "Name | Status | Start Date | End Date | Progress")
            y -= 20
            for proj in data['projects']:
                line = f"{proj['name']} | {proj['status']} | {proj['start_date']} | {proj['end_date']} | {proj['progress']}%"
                p.drawString(50, y, line)
                y -= 15
                if y < 50:
                    p.showPage()
                    y = height - 50
        elif report_type == 'task':
            p.drawString(50, y, "Title | Status | Priority | Due Date | Project | Assignee")
            y -= 20
            for t in data.get('overdue_tasks', []):
                line = f"{t['title']} | {t['status']} | {t.get('priority', '')} | {t['due_date']} | {t.get('project__name', '')} | {t.get('assigned_to__username', '')}"
                p.drawString(50, y, line)
                y -= 15
                if y < 50:
                    p.showPage()
                    y = height - 50
        elif report_type == 'team':
            for team in data['teams']:
                p.drawString(50, y, f"Team: {team['name']} (Type: {team['type']}, Manager: {team['manager']})")
                y -= 15
                p.drawString(60, y, "Members: Name | Role | Completed | Pending | Hours")
                y -= 15
                for m in team['members']:
                    line = f"{m['name']} | {m['role']} | {m['completed_tasks']} | {m['pending_tasks']} | {m['total_hours']}"
                    p.drawString(60, y, line)
                    y -= 13
                    if y < 50:
                        p.showPage()
                        y = height - 50
                y -= 10
        elif report_type == 'time':
            p.drawString(50, y, "User | Project | Hours | Date | Description")
            y -= 20
            for entry in data['time_entries']:
                line = f"{entry['user']} | {entry['project']} | {entry['hours']} | {entry['date']} | {entry['description']}"
                p.drawString(50, y, line)
                y -= 15
                if y < 50:
                    p.showPage()
                    y = height - 50
        elif report_type == 'burndown':
            p.drawString(50, y, "Date | Remaining | Completed")
            y -= 20
            for point in data['series']:
                line = f"{point['date']} | {point['remaining']} | {point['completed']}"
                p.drawString(50, y, line)
                y -= 15
                if y < 50:
                    p.showPage()
                    y = height - 50
        elif report_type == 'activity':
            p.drawString(50, y, "Notifications:")
            y -= 15
            for n in data['notifications']:
                line = f"{n['timestamp']} | {n['user']} | {n['type']} | {n['title']} | {n['message']}"
                p.drawString(50, y, line)
                y -= 13
                if y < 50:
                    p.showPage()
                    y = height - 50
            y -= 10
            p.drawString(50, y, "Comments:")
            y -= 15
            for c in data['comments']:
                line = f"{c['timestamp']} | {c['user']} | {c.get('project', '')} | {c.get('task', '')} | {c['content']}"
                p.drawString(50, y, line)
                y -= 13
                if y < 50:
                    p.showPage()
                    y = height - 50
        p.save()
        buffer.seek(0)
        from django.http import HttpResponse
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.pdf"'
        return response

    def generate_excel(self, report_type, data):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.capitalize()} Report"
        if report_type == 'project':
            ws.append(["Name", "Status", "Start Date", "End Date", "Progress"])
            for proj in data['projects']:
                ws.append([
                    proj['name'],
                    proj['status'],
                    proj['start_date'],
                    proj['end_date'],
                    proj['progress']
                ])
        elif report_type == 'task':
            ws.append(["Title", "Status", "Priority", "Due Date", "Project", "Assignee"])
            for t in data.get('overdue_tasks', []):
                ws.append([
                    t['title'],
                    t['status'],
                    t.get('priority', ''),
                    t['due_date'],
                    t.get('project__name', ''),
                    t.get('assigned_to__username', '')
                ])
        elif report_type == 'team':
            ws.append(["Team", "Type", "Manager", "Member Name", "Role", "Completed Tasks", "Pending Tasks", "Total Hours"])
            for team in data['teams']:
                for m in team['members']:
                    ws.append([
                        team['name'],
                        team['type'],
                        team['manager'],
                        m['name'],
                        m['role'],
                        m['completed_tasks'],
                        m['pending_tasks'],
                        m['total_hours']
                    ])
        elif report_type == 'time':
            ws.append(["User", "Project", "Hours", "Date", "Description"])
            for entry in data['time_entries']:
                ws.append([
                    entry['user'],
                    entry['project'],
                    entry['hours'],
                    entry['date'],
                    entry['description']
                ])
        elif report_type == 'burndown':
            ws.append(["Date", "Remaining", "Completed"])
            for point in data['series']:
                ws.append([point['date'], point['remaining'], point['completed']])
        elif report_type == 'activity':
            ws.append(["Type", "Timestamp", "User", "Title", "Message"])
            for n in data['notifications']:
                ws.append([
                    n['type'],
                    n['timestamp'],
                    n['user'],
                    n['title'],
                    n['message']
                ])
            ws.append([])
            ws.append(["Comment Timestamp", "User", "Project", "Task", "Content"])
            for c in data['comments']:
                ws.append([
                    c['timestamp'],
                    c['user'],
                    c.get('project', ''),
                    c.get('task', ''),
                    c['content']
                ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        from django.http import HttpResponse
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
        return response




class LogoutView(APIView):
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        print("logout in post")
        try:
            user_to_log = request.user
            # Invalidate session (for session authentication)
            logout(request)
            role = user_to_log.userprofile.role if hasattr(user_to_log, 'userprofile') else 'User'
            role_cleaned = role.replace(" ", "_").upper()

            log_type = f"{role_cleaned}_LOGGED_OUT"
            message = f"{role} {user_to_log.username} logged out successfully."

            create_activity_log(user_to_log, log_type, message)

            # Get the refresh token from the request data (for JWT token authentication)
            refresh_token = request.data.get('refresh_token')
            print("refreshtoken",refresh_token)
            if not refresh_token:
                return JsonResponse({'detail': 'Refresh token not provided'}, status=status.HTTP_400_BAD_REQUEST)
            return JsonResponse({'detail': 'Logout successful'}, status=status.HTTP_205_RESET_CONTENT)
        except AuthenticationFailed:
            return JsonResponse({'detail': 'Authentication failed'}, status=status.HTTP_401_UNAUTHORIZED)
        except Exception as e:
            return JsonResponse({'detail': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@login_required
def get_chat_room_and_messages(request):
    user = request.user
    employee_username = request.GET.get("Employee")

    if not employee_username:
        return JsonResponse({"error": "Employee not specified"}, status=400)

    try:
        other_user = User.objects.get(username=employee_username)
    except User.DoesNotExist:
        return JsonResponse({"error": "Employee not found"}, status=404)

    usernames = sorted([user.username, other_user.username])
    room_name = f"{usernames[0]}_{usernames[1]}"

    room, created = Room.objects.get_or_create(name=room_name)
    if created:
        room.room_type = 'private'
        room.save()

    messages = ChatMessage.objects.filter(room=room).select_related('user').order_by('-timestamp')[:50]

    return JsonResponse({
        "room_name": room_name,
        "messages": [
            {
                'username': msg.user.username,
                'message': msg.message,
                'timestamp': msg.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            } for msg in reversed(messages)
        ]
    })


@login_required
def get_public_messages(request):
    room_name = "Group_Avsys"
    room, created = Room.objects.get_or_create(name=room_name)
    if created:
        room.room_type = 'public'
        room.save()

    messages = ChatMessage.objects.filter(room=room).select_related('user').order_by('-timestamp')[:50]

    return JsonResponse({
        "messages": [
            {
                'username': msg.user.username,
                'message': msg.message,
                'timestamp': msg.timestamp.strftime('%Y-%m-%d %H:%M:%S')
            } for msg in reversed(messages)
        ]
    })

# Helper function for logging with deduplication
def create_activity_log(user, action_type, description, content_object=None):
    # Avoid logging for anonymous users if not desired, but here we allow it for actions like registration
    if not user.is_authenticated:
        user = None

    # Check for duplicate entries within the last 5 seconds to prevent spam
    from django.utils import timezone
    from datetime import timedelta
    
    recent_time = timezone.now() - timedelta(seconds=5)
    
    # Check if a similar log entry was created recently
    existing_log = ActivityLog.objects.filter(
        user=user,
        action_type=action_type,
        description=description,
        timestamp__gte=recent_time
    ).first()
    
    if existing_log:
        # If duplicate found within 5 seconds, don't create another entry
        return existing_log

    log_entry = ActivityLog(
        user=user,
        action_type=action_type,
        description=description,
    )
    if content_object:
        log_entry.content_type = ContentType.objects.get_for_model(content_object)
        log_entry.object_id = content_object.pk
    log_entry.save()
    return log_entry


from django.utils.dateparse import parse_date

class LogsView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        
        role = getattr(request.user.userprofile, 'role', 'User')
        role_cleaned = role.replace(" ", "_").upper()
        create_activity_log(
            request.user,
            f"{role_cleaned}_LOGS_PAGE_VISIT",
            f"{role.title()} {request.user.username} visited the Logs page."
        )

        profile = getattr(request.user, 'userprofile', None)
        if not (request.user.is_superuser or (profile and profile.role in ['Manager', 'Team Lead','Developer', 'Tester'])):
            messages.error(request, "You are not authorized to view this page.")
            return redirect('dashboard')

        # Base queryset
        if request.user.is_superuser or profile.role in ['Manager', 'Team Lead']:
            logs = ActivityLog.objects.all().select_related('user')
        else:
            logs = ActivityLog.objects.filter(user=request.user).select_related('user')

        # Server-side filters from GET
        search = request.GET.get('search', '')
        from_date = request.GET.get('from_date')
        to_date = request.GET.get('to_date')
        action_type = request.GET.get('action', '')

        if search:
            logs = logs.filter(
                Q(description__icontains=search) |
                Q(action_type__icontains=search) |
                Q(user__username__icontains=search)
            )
        if from_date:
            logs = logs.filter(timestamp__date__gte=parse_date(from_date))
        if to_date:
            logs = logs.filter(timestamp__date__lte=parse_date(to_date))
        if action_type:
            logs = logs.filter(action_type=action_type)

        logs = logs.order_by('-timestamp')
        return render(request, 'logs.html', {'logs': logs})
    

  
class AddCommentView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
        
        try:
            data = request.data
            content = data.get('content')
            task_id = data.get('task_id')
            project_id = data.get('project_id')
            
            if not content:
                return Response({'error': 'Comment content is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Create comment
            comment = Comment.objects.create(
                user=request.user,
                content=content,
                task_id=task_id if task_id else None,
                project_id=project_id if project_id else None
            )
            
            # Notify manager if assignee is commenting on their assigned task
            if task_id:
                try:
                    task = Task.objects.get(id=task_id)
                    # Check if the commenter is the assigned user and not a manager
                    if (task.assigned_to == request.user and 
                        request.user.userprofile.role not in ['Manager', 'Team Lead']):
                        # Find the manager who created the task or is the project creator
                        manager = task.created_by
                        if manager and manager != request.user:
                            create_notification(
                                user=manager,
                                title="Task Comment",
                                message=f"{request.user.username} commented on task: {task.title}",
                                notification_type="Task Updated",
                                link=f"/tasks/{task.id}/view/"
                            )
                except Task.DoesNotExist:
                    pass
            
            # Log the comment creation
            role = getattr(request.user.userprofile, 'role', 'User')
            role_cleaned = role.replace(" ", "_").upper()
            
            log_type = f"{role_cleaned}_COMMENT_CREATED"
            message = f'{role.title()} {request.user.username} added a comment.'
            
            create_activity_log(request.user, log_type, message, comment)
            
            return Response({
                'success': True,
                'message': 'Comment added successfully',
                'comment': {
                    'id': comment.id,
                    'content': comment.content,
                    'user': comment.user.username,
                    'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M')
                }
            })
            
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
class DailyStatusListCreateView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        """
        Handle GET requests:
        - If normal browser request: render HTML page
        - If AJAX request: return JSON of statuses, filtered by role and members
        """
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('as_json')
        user = request.user
        profile = user.userprofile

        # 🌐 Normal page load
        if not is_ajax:
            role_cleaned = profile.role.replace(" ", "_").upper()
            create_activity_log(
                user,
                f"{role_cleaned}_DAILY_STATUS_PAGE_VISIT",
                f"{profile.role} {user.username} visited the Daily Status page."
            )

            # Filter projects for developer
            if profile.role == 'Developer':
                projects = Project.objects.filter(teams__members=user).distinct()
            else:
                projects = Project.objects.all()

            return render(request, 'daily_status.html', {'projects': projects})

        # 🌐 AJAX / JSON request
        team_id = request.GET.get('team')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        project_id = request.GET.get('project_id')

        # ✅ New: multiple member IDs from query (?member_ids=1&member_ids=2)
        member_ids = request.GET.getlist('member_ids')

        queryset = DailyStatus.objects.select_related('user', 'project', 'team')

        # Role-based visibility
        if profile.role == 'Manager' and team_id:
            queryset = queryset.filter(team_id=team_id)
        elif profile.role == 'Developer':
            user_projects = Project.objects.filter(teams__members=user).distinct()
            queryset = queryset.filter(project__in=user_projects, user=user)
        else:
            queryset = queryset.filter(user=user)

        # ✅ New: filter by project
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        # ✅ New: filter by multiple members
        if member_ids:
            queryset = queryset.filter(user_id__in=member_ids)

        # ✅ Date range
        if start_date and end_date:
            queryset = queryset.filter(date__range=[start_date, end_date])

        data = [{
            'id': status.id,
            'user_id': status.user.id,
            'user_name': status.user.username,
            'team': status.team.team_name if status.team else '',
            'project_id': status.project.id if status.project else None,
            'project_name': status.project.name if status.project else '',
            'date': status.date.strftime('%Y-%m-%d'),
            'status': status.status_text,
        } for status in queryset.order_by('-date')]

        return JsonResponse({'statuses': data})


    def post(self, request):
        """
        Handle POST requests to create a new daily status
        """
        try:
            data = json.loads(request.body)
            status_text = data.get('status_text')
            date_str = data.get('date', datetime.now().date().isoformat())
            project_id = data.get('project_id')

            if not all([status_text, project_id]):
                return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)

            user = request.user
            profile = user.userprofile

            # Restrict non-managers to their own projects (via Teams.members or ProjectMember or direct members)
            if profile.role not in ['Manager', 'Team Lead']:
                try:
                    pid = int(project_id)
                except Exception:
                    return JsonResponse({'success': False, 'error': 'Invalid project'}, status=400)
                allowed_projects = Project.objects.filter(
                    Q(teams__members=user) |
                    Q(project_members__user=user) |
                    Q(members=user)
                ).values_list('id', flat=True).distinct()
                if pid not in allowed_projects:
                    return JsonResponse({'success': False, 'error': 'You cannot submit status for a project you are not assigned to'}, status=403)

            # Convert date string to date object
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date() if isinstance(date_str, str) else date_str

            status = DailyStatus.objects.create(
                user=user,
                project_id=project_id,
                date=date_obj,
                status_text=status_text
            )

            return JsonResponse({
                'success': True,
                'status': {
                    'id': status.id,
                    'date': status.date.strftime('%Y-%m-%d'),
                    'status_text': status.status_text,
                    'project_id': status.project_id,
                    'user_id': status.user_id
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    def put(self, request, status_id=None):
        try:
            data = json.loads(request.body)
            # Fallback: get id from body if not passed in URL
            status_id = status_id or data.get('id')

            if not status_id:
                return JsonResponse({'success': False, 'error': 'Status ID missing'}, status=400)

            status = DailyStatus.objects.get(id=status_id)
            user = request.user
            profile = user.userprofile

            # Only allow manager/team lead or owner to edit
            if user != status.user and profile.role not in ['Manager', 'Team Lead']:
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

            project_id = data.get('project_id')
            status_text = data.get('status_text')

            # Non-managers can only assign/update projects they belong to
            if profile.role not in ['Manager', 'Team Lead']:
                if project_id:
                    try:
                        pid = int(project_id)
                    except Exception:
                        return JsonResponse({'success': False, 'error': 'Invalid project'}, status=400)
                    allowed_projects = Project.objects.filter(
                        Q(teams__members=user) |
                        Q(project_members__user=user) |
                        Q(members=user)
                    ).values_list('id', flat=True).distinct()
                    if pid not in allowed_projects:
                        return JsonResponse({'success': False, 'error': 'Cannot update project you are not assigned to'}, status=403)

            if project_id:
                status.project_id = project_id
            if status_text:
                status.status_text = status_text
            status.save()

            return JsonResponse({
                'success': True,
                'status': {
                    'id': status.id,
                    'project_id': status.project_id,
                    'user_id': status.user_id,
                    'status_text': status.status_text,
                    'date': status.date.strftime('%Y-%m-%d')
                }
            })

        except DailyStatus.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Status not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


class DailyStatusExportView(View):
    def get(self, request, user_id):
        # Only managers can export
        if not request.user.userprofile.role == 'Manager':
            return HttpResponseForbidden("Not allowed")
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        statuses = DailyStatus.objects.filter(user_id=user_id)
        if start_date and end_date:
            statuses = statuses.filter(date__range=[start_date, end_date])

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Status"
        ws.append(['Date', 'Status'])

        for status in statuses.order_by('date'):
            ws.append([status.date.strftime('%Y-%m-%d'), status.status_text])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=daily_status_{user_id}.xlsx'
        wb.save(response)
        return response

class DailyStatusTeamExportView(View):
    def get(self, request, team_id):
        if not request.user.userprofile.role == 'Manager':
            return HttpResponseForbidden("Not allowed")
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        team = Teams.objects.get(id=team_id)
        members = User.objects.filter(userprofile__team=team)
        wb = Workbook()
        for idx, member in enumerate(members):
            if idx == 0:
                ws = wb.active
                ws.title = member.username
            else:
                ws = wb.create_sheet(title=member.username)
            statuses = DailyStatus.objects.filter(user=member)
            if start_date and end_date:
                statuses = statuses.filter(date__range=[start_date, end_date])
            ws.append(['Date', 'Status'])
            for status in statuses.order_by('date'):
                ws.append([status.date.strftime('%Y-%m-%d'), status.status_text])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=team_{team_id}_daily_status.xlsx'
        wb.save(response)
        return response

class DailyStatusTemplateView(View):
    def get(self, request):
        # Only managers can generate template
        if not request.user.userprofile.role == 'Manager':
            return HttpResponseForbidden("Not allowed")
        team_id = request.GET.get('team_id')
        member_ids = request.GET.getlist('member_ids')  # e.g., ?member_ids=1&member_ids=2
        members = []
        if team_id:
            team = Teams.objects.get(id=team_id)
            members = User.objects.filter(userprofile__team=team)
        elif member_ids:
            members = User.objects.filter(id__in=member_ids)
        else:
            # If nothing selected, return all employees
            members = User.objects.filter(userprofile__role__in=['Developer', 'Tester', 'Designer'])
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Status Template"
        ws.append(['Date', 'Employee Name', 'Status'])
        for member in members:
            ws.append(['', member.username, ''])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=daily_status_template.xlsx'
        wb.save(response)
        return response
  


class ProjectListAPI(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return JsonResponse({'projects': []})

        try:
            profile = request.user.userprofile
            role = getattr(profile, 'role', 'User')
        except Exception:
            role = 'User'

        queryset = Project.objects.all()

        # Restrict based on role
        if role in ['Developer', 'Tester', 'Designer']:
            # Projects where the user is associated via Teams.members OR ProjectMember OR direct members
            queryset = queryset.filter(
                Q(teams__members=request.user) |
                Q(project_members__user=request.user) |
                Q(members=request.user)
            ).distinct()

        projects = queryset.values('id', 'name')
        return JsonResponse({'projects': list(projects)})

class TeamListAPI(View):
    def get(self, request, project_id):
        # Optionally, filter teams by project if you have such a relation
        teams = Teams.objects.all().values('id', 'team_name', 'team_type')
        return JsonResponse({'teams': list(teams)})

class MemberListAPI(View):
    def get(self, request, team_id):
        members = UserProfile.objects.filter(team_id=team_id).select_related('user')
        data = [{'id': m.user.id, 'name': m.user.username} for m in members]
        return JsonResponse({'members': data})

class ProjectMembersAPI(View):
    def get(self, request, project_id):
        """
        Return all users who are members of the given project.
        Sources:
        - Project.members (M2M)
        - ProjectMember (explicit membership with role)
        - Teams.members where Teams.project = project
        """
        try:
            user_ids = set()

            # Project.members M2M
            try:
                project = Project.objects.get(id=project_id)
                user_ids.update(project.members.values_list('id', flat=True))
            except Project.DoesNotExist:
                return JsonResponse({'members': []})

            # ProjectMember relations (also capture roles)
            pm_qs = ProjectMember.objects.filter(project_id=project_id).values('user_id', 'role')
            user_role_map = {}
            for row in pm_qs:
                uid = row['user_id']
                role = row['role']
                # Handle JSONField role - it could be a string or list
                if isinstance(role, list) and role:
                    role_str = ', '.join(role)
                elif isinstance(role, str) and role:
                    role_str = role
                else:
                    role_str = 'Member'
                user_ids.add(uid)
                user_role_map.setdefault(uid, set()).add(role_str)

            # Teams.members for this project (role may not be known here)
            team_member_ids = Teams.objects.filter(project_id=project_id).values_list('members__id', flat=True)
            for uid in team_member_ids:
                if uid:
                    user_ids.add(uid)

            # Build response including role (join multiple roles if present)
            users = User.objects.filter(id__in=user_ids).order_by('username')
            data = []
            for u in users:
                roles = user_role_map.get(u.id)
                role_label = ', '.join(sorted(roles)) if roles else 'Member'
                data.append({'id': u.id, 'name': u.username, 'role': role_label})
            
            return JsonResponse({'members': data})
        except Exception as e:
            return JsonResponse({'members': [], 'error': str(e)}, status=500)

class ProjectPageNamesAPI(View):
    def get(self, request, project_id):
        """
        Return distinct page names used by tasks under the given project.
        """
        try:
            names = Task.objects.filter(project_id=project_id) \
                .exclude(page_name__isnull=True) \
                .exclude(page_name__exact='') \
                .values_list('page_name', flat=True).distinct().order_by('page_name')
            return JsonResponse({'page_names': list(names)})
        except Exception as e:
            return JsonResponse({'page_names': [], 'error': str(e)}, status=500)
class StatusListAPI(View):
    def get(self, request):
        user_ids = request.GET.getlist('user_ids') or []
        project_id = request.GET.get('project_id')
        start = request.GET.get('start')
        end = request.GET.get('end')
        
        # Require authentication; otherwise return empty
        if not request.user.is_authenticated:
            return JsonResponse({'statuses': []})

        qs = DailyStatus.objects.select_related('user', 'project').all()

        # Role-based visibility (same logic as before)
        try:
            role = request.user.userprofile.role
        except Exception:
            role = 'User'

        if role in ['Manager', 'Team Lead']:
            pass  # can see all
        else:
            # Non-managers see only projects they belong to
            user_projects = Project.objects.filter(
                Q(teams__members=request.user) |
                Q(project_members__user=request.user) |
                Q(members=request.user)
            ).distinct()
            qs = qs.filter(project__in=user_projects)

        # ✅ Filter by multiple user IDs
        if user_ids:
            qs = qs.filter(user_id__in=user_ids)

        # Single user_id fallback
        user_id = request.GET.get('user_id')
        if user_id and not user_ids:
            qs = qs.filter(user_id=user_id)

        if project_id:
            qs = qs.filter(project_id=project_id)
        if start and end:
            qs = qs.filter(date__range=[start, end])
            
        data = [{
            'id': s.id,
            'date': s.date.strftime('%Y-%m-%d'),
            'status': s.status_text,
            'project_id': s.project_id,
            'user_id': s.user_id,
            'user_name': s.user.username,
            'project_name': s.project.name if s.project else ''
        } for s in qs.order_by('date')]
        
        return JsonResponse({'statuses': data})

class AllMembersAPI(View):
    def get(self, request):
        users = User.objects.all().select_related('userprofile').order_by('username')
        data = []
        for user in users:
            try:
                profile = user.userprofile
                name = profile.name or user.username
                role = profile.role or 'Member'
            except Exception:
                name = user.username
                role = 'Member'
            data.append({
                'id': user.id,
                'name': name,
                'role': role
            })
        return JsonResponse({'members': data})
